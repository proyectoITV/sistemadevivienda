# -*- coding: utf-8 -*-
import json
import time
from datetime import timedelta, datetime
from urllib import parse, request as urllib_request

from django.contrib import messages
from django.conf import settings
from django.shortcuts import redirect, render, get_object_or_404
from django.utils import timezone
from django.db import ProgrammingError, OperationalError, transaction, IntegrityError
from django.db.models import Q
from django.core.exceptions import ObjectDoesNotExist
from django.core.paginator import Paginator
from django.core.files.storage import default_storage
from django.http import JsonResponse
from django.contrib.auth.decorators import login_required
from django.http import JsonResponse, FileResponse
from .models import Anuncio, SeguridadContactanos, CatalogosDelegaciones
from .email_utils import enviar_correo_contacto

def index(request):
	if request.user.is_authenticated:
		return redirect('dashboard')

	anuncios = Anuncio.objects.filter(activo=True).order_by('-fecha_publicacion')
	return render(request, 'desarrollo/web/index.html', {'anuncios': anuncios})


def _verify_recaptcha(token, remote_ip):
	secret = getattr(settings, 'RECAPTCHA_SECRET_KEY', '')
	if not secret:
		return True

	if not token:
		return False

	data = parse.urlencode({
		'secret': secret,
		'response': token,
		'remoteip': remote_ip or '',
	}).encode()
	req = urllib_request.Request('https://www.google.com/recaptcha/api/siteverify', data=data)

	try:
		with urllib_request.urlopen(req, timeout=8) as response:
			result = json.loads(response.read().decode('utf-8'))
			return bool(result.get('success'))
	except Exception:
		return False


def contactanos(request):
	min_seconds = getattr(settings, 'CONTACT_FORM_MIN_SECONDS', 3)
	rate_limit_seconds = getattr(settings, 'CONTACT_FORM_RATE_LIMIT_SECONDS', 60)
	recaptcha_site_key = getattr(settings, 'RECAPTCHA_SITE_KEY', '')
	uuid_folio = None

	if request.method == 'POST':
		form = SeguridadContactanosForm(request.POST)
		start_at = request.session.get('contact_form_started_at')
		elapsed = time.time() - float(start_at) if start_at else 0

		if start_at and elapsed < min_seconds:
			messages.error(request, 'Envío demasiado rápido. Intenta nuevamente en unos segundos.')
		elif form.is_valid():
			ip_origen = request.META.get('REMOTE_ADDR')
			window_start = timezone.now() - timedelta(seconds=rate_limit_seconds)

			if ip_origen and SeguridadContactanos.objects.filter(ip_origen=ip_origen, fecha_hora_insercion__gte=window_start).exists():
				messages.error(request, 'Ya recibimos un mensaje reciente desde tu conexión. Intenta nuevamente en un minuto.')
			else:
				recaptcha_token = request.POST.get('g-recaptcha-response', '')
				if not _verify_recaptcha(recaptcha_token, ip_origen):
					messages.error(request, 'No se pudo validar reCAPTCHA. Verifica e inténtalo de nuevo.')
				else:
					contacto = form.save(commit=False)
					contacto.ip_origen = ip_origen
					contacto.user_agent = request.META.get('HTTP_USER_AGENT', '')
					contacto.save()
					
					# Enviar correo de confirmación
					enviar_correo_contacto(contacto)
					
					uuid_folio = str(contacto.uuid_folio)
					request.session['contact_form_started_at'] = time.time()
					request.session['uuid_folio'] = uuid_folio
					return redirect('contactanos')

		request.session['contact_form_started_at'] = time.time()
	else:
		form = SeguridadContactanosForm()
		request.session['contact_form_started_at'] = time.time()
		uuid_folio = request.session.pop('uuid_folio', None)

	return render(request, 'desarrollo/web/contactanos.html', {
		'form': form, 
		'recaptcha_site_key': recaptcha_site_key,
		'uuid_folio': uuid_folio
	})


def buscar_seguimiento(request):
	contacto = None
	busqueda = request.GET.get('folio', '').strip()

	if busqueda:
		try:
			contacto = SeguridadContactanos.objects.get(uuid_folio=busqueda)
		except SeguridadContactanos.DoesNotExist:
			messages.error(request, 'Folio no encontrado. Verifica e intenta nuevamente.')

	return render(request, 'desarrollo/web/seguimiento.html', {'contacto': contacto, 'busqueda': busqueda})


def nuestras_oficinas(request):
	delegaciones = CatalogosDelegaciones.objects.filter(activo=True).order_by('nombre')
	return render(request, 'desarrollo/web/nuestras_oficinas.html', {'delegaciones': delegaciones})

# ============== AUTENTICACIÓN ==============

from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required
from django.contrib.auth.models import AnonymousUser
from .models import PersonalEmpleados, RecuperacionContrasena, PersonalDepartamento, TransparenciaGo
from .email_utils import enviar_correo_recuperacion


def _usuario_es_admin_sistema(user):
	"""Valida si el usuario autenticado tiene perfil de usuario_sistema con rol admin."""
	if not getattr(user, 'is_authenticated', False):
		return False

	# Compatibilidad con usuarios administradores del modelo principal
	if getattr(user, 'is_superuser', False) or getattr(user, 'is_staff', False) or getattr(user, 'is_admin', False):
		return True

	try:
		usuario_sistema = user.usuario_sistema
		return usuario_sistema.activo and usuario_sistema.rol == 'admin'
	except ObjectDoesNotExist:
		return False


def login_view(request):
	"""Vista para iniciar sesión"""
	if request.user.is_authenticated:
		return redirect('dashboard')
	
	# Mostrar mensaje si la sesión fue cerrada por inactividad
	if request.session.pop('sesion_expirada_por_inactividad', False):
		messages.warning(request, 
			'Tu sesión ha sido cerrada por inactividad. Por favor, inicia sesión nuevamente.'
		)
	
	if request.method == 'POST':
		usuario = request.POST.get('usuario', '').strip()
		password = request.POST.get('password', '').strip()
		recordar = request.POST.get('recordar_dispositivo', False)
		
		if not usuario or not password:
			messages.error(request, 'Por favor ingresa usuario y contraseña.')
			return render(request, 'desarrollo/web/login.html')
		
		# Intentar autenticar con usuario o email
		user = None
		try:
			personal = PersonalEmpleados.objects.get(usuario=usuario)
			if personal.check_password(password):
				user = personal
		except PersonalEmpleados.DoesNotExist:
			# Intentar con email
			try:
				personal = PersonalEmpleados.objects.get(email=usuario)
				if personal.check_password(password):
					user = personal
			except PersonalEmpleados.DoesNotExist:
				pass
		
		if user and user.is_active:
			login(request, user)
			user.fecha_ultimo_login = timezone.now()
			
			# Guardar token del dispositivo si lo solicita
			if recordar:
				from django.utils.crypto import get_random_string
				user.token_dispositivo = get_random_string(64)
				user.recordar_dispositivo = True
			
			user.save()
			
			# Inicializar timestamp de última actividad para el middleware
			request.session['ultima_actividad'] = timezone.now().timestamp()
			
			messages.success(request, f'¡Bienvenido {user.nombre_completo}!')
			return redirect('intro')
		else:
			messages.error(request, 'Usuario, email o contraseña incorrectos.')
	
	return render(request, 'desarrollo/web/login.html')


def logout_view(request):
	"""Vista para cerrar sesión"""
	logout(request)
	messages.success(request, 'Sesión cerrada correctamente.')
	return redirect('index')


@login_required(login_url='login')
def intro(request):
	"""Vista de introducción después del login"""
	from .models import ConfiguracionSistema
	
	# Obtener configuración del sistema
	config = ConfiguracionSistema.objects.first()
	duracion = config.duracion_intro_segundos if config else 3
	
	context = {
		'duracion_intro': duracion,
	}
	return render(request, 'desarrollo/web/intro.html', context)


@login_required(login_url='login')
def dashboard(request):
	"""Dashboard principal del sistema"""
	# Si no hay sesión activa, redirigir al index
	if not request.user.is_authenticated:
		return redirect('index')

	# Indicadores rápidos
	total_empleados = PersonalEmpleados.objects.filter(activo=True).count()

	# Conteo de mantenimiento visible para el usuario
	es_admin = _usuario_es_admin_sistema(request.user)
	q_mantenimiento_recibidos = Q(pk__in=[])
	if es_admin:
		q_mantenimiento_recibidos = Q()
	elif request.user.iddepartamento_id:
		q_mantenimiento_recibidos = Q(departamento_soporte_id=request.user.iddepartamento_id)

	qs_mantenimiento = TicketMantenimiento.objects.filter(
		Q(solicitante=request.user) | q_mantenimiento_recibidos
	)
	if es_admin:
		qs_mantenimiento = TicketMantenimiento.objects.all()

	conteo_mantenimiento_total = qs_mantenimiento.count()
	conteo_mantenimiento_pendiente = qs_mantenimiento.filter(
		estado__in=['abierto', 'asignado', 'en_revision', 'en_reparacion', 'espera_refaccion']
	).count()
	conteo_mantenimiento_vencido = qs_mantenimiento.filter(
		estado__in=['abierto', 'asignado', 'en_revision', 'en_reparacion', 'espera_refaccion'],
		fecha_limite_sla__lt=timezone.now()
	).count()

	context = {
		'total_empleados': total_empleados,
		'conteo_mantenimiento_total': conteo_mantenimiento_total,
		'conteo_mantenimiento_pendiente': conteo_mantenimiento_pendiente,
		'conteo_mantenimiento_vencido': conteo_mantenimiento_vencido,
	}
	return render(request, 'desarrollo/web/dashboard.html', context)


@login_required(login_url='login')
def listar_archivos_transparencia(request):
	"""Listado de archivos de transparencia con carga de nuevos PDF."""
	if request.method == 'POST':
		form = TransparenciaArchivoUploadForm(request.POST, request.FILES)
		if form.is_valid():
			with transaction.atomic():
				ahora = timezone.localtime(timezone.now())
				usuario_carga = getattr(request.user, 'usuario', '') or getattr(request.user, 'username', '') or 'sistema'
				ultimo_registro = TransparenciaGo.objects.select_for_update().order_by('-id_file').first()
				nuevo_id = (ultimo_registro.id_file if ultimo_registro else 0) + 1

				# 1) Registro en BD con IdFile incremental (ultimo + 1)
				registro = TransparenciaGo.objects.create(
					id_file=nuevo_id,
					file_nombre=form.cleaned_data['nombre'].strip(),
					id_user=usuario_carga,
					fecha=ahora.date(),
					hora=ahora.time().replace(microsecond=0),
					file_descripcion=form.cleaned_data.get('comentarios', '').strip(),
				)

				# 2) Renombrado dinamico usando el IdFile: Transparencia/Files/{IdFile}.pdf
				ruta_pdf = f"Transparencia/Files/{registro.id_file}.pdf"
				if default_storage.exists(ruta_pdf):
					default_storage.delete(ruta_pdf)
				default_storage.save(ruta_pdf, form.cleaned_data['archivo_pdf'])

			messages.success(request, 'Archivo de transparencia cargado correctamente.')
			return redirect('listar_archivos_transparencia')
		messages.error(request, 'No fue posible cargar el archivo. Verifica los campos e intenta de nuevo.')
	else:
		form = TransparenciaArchivoUploadForm()

	archivos_qs = TransparenciaGo.objects.all().order_by('-fecha', '-hora', '-id_file')
	paginator = Paginator(archivos_qs, 10)
	page_number = request.GET.get('page')
	page_obj = paginator.get_page(page_number)

	for archivo in page_obj.object_list:
		fecha_hora_registro = datetime.combine(archivo.fecha, archivo.hora)
		if timezone.is_naive(fecha_hora_registro):
			fecha_hora_registro = timezone.make_aware(fecha_hora_registro, timezone.get_current_timezone())
		tiempo_transcurrido = timezone.now() - fecha_hora_registro
		archivo.puede_eliminar = tiempo_transcurrido < timedelta(hours=24)
		limite_eliminacion = timedelta(hours=24)

		if archivo.puede_eliminar:
			tiempo_restante = limite_eliminacion - tiempo_transcurrido
			segundos_restantes = max(0, int(tiempo_restante.total_seconds()))
			horas_restantes = segundos_restantes // 3600
			minutos_restantes = (segundos_restantes % 3600) // 60
			archivo.tooltip_eliminar = (
				f"Se puede eliminar. Tiempo restante: {horas_restantes:02d}h {minutos_restantes:02d}m"
			)
		else:
			tiempo_bloqueado = tiempo_transcurrido - limite_eliminacion
			segundos_bloqueado = max(0, int(tiempo_bloqueado.total_seconds()))
			horas_bloqueado = segundos_bloqueado // 3600
			minutos_bloqueado = (segundos_bloqueado % 3600) // 60
			archivo.tooltip_eliminar = (
				f"Eliminacion bloqueada (>=24h). Bloqueado desde hace {horas_bloqueado:02d}h {minutos_bloqueado:02d}m"
			)

		ruta_pdf = f"Transparencia/Files/{archivo.id_file}.pdf"
		archivo.pdf_url = f"{settings.MEDIA_URL}{ruta_pdf}" if default_storage.exists(ruta_pdf) else ''
		archivo.ruta_directa = ruta_pdf

	context = {
		'page_obj': page_obj,
		'form_upload': form,
		'total_archivos': archivos_qs.count(),
	}
	return render(request, 'desarrollo/Transparencia/listar_archivos.html', context)


@login_required(login_url='login')
def eliminar_archivo_transparencia(request, id_file):
	"""Elimina archivo de transparencia solo si tiene menos de 24 horas desde su registro."""
	if request.method != 'POST':
		messages.error(request, 'Metodo no permitido para eliminar archivos.')
		return redirect('listar_archivos_transparencia')

	archivo = get_object_or_404(TransparenciaGo, id_file=id_file)
	fecha_hora_registro = datetime.combine(archivo.fecha, archivo.hora)
	if timezone.is_naive(fecha_hora_registro):
		fecha_hora_registro = timezone.make_aware(fecha_hora_registro, timezone.get_current_timezone())

	tiempo_transcurrido = timezone.now() - fecha_hora_registro
	if tiempo_transcurrido >= timedelta(hours=24):
		messages.warning(request, 'No se puede eliminar: el archivo tiene 24 horas o mas de antiguedad.')
		return redirect('listar_archivos_transparencia')

	ruta_pdf = f"Transparencia/Files/{archivo.id_file}.pdf"
	with transaction.atomic():
		archivo.delete()
		if default_storage.exists(ruta_pdf):
			default_storage.delete(ruta_pdf)

	messages.success(request, 'Archivo eliminado correctamente.')
	return redirect('listar_archivos_transparencia')


def recuperar_contrasena(request):
	"""Vista para solicitar recuperación de contraseña"""
	if request.method == 'POST':
		usuario_o_email = request.POST.get('usuario_email', '').strip()
		
		if not usuario_o_email:
			messages.error(request, 'Por favor ingresa tu usuario o correo.')
			return render(request, 'desarrollo/web/recuperar_contrasena.html')
		
		# Buscar el usuario por usuario o email
		try:
			usuario = PersonalEmpleados.objects.get(
				Q(usuario=usuario_o_email) | Q(email=usuario_o_email)
			)
			
			# Crear token de recuperación
			recuperacion = RecuperacionContrasena.crear_token_recuperacion(
				usuario, 
				ip_origen=get_client_ip(request)
			)
			
			# Enviar correo
			enviar_correo_recuperacion(usuario, recuperacion)
			
			messages.success(request, 
				f'Se ha enviado un correo de recuperación a {usuario.email}. '
				'El enlace será válido por 24 horas.'
			)
			return redirect('login')
			
		except PersonalEmpleados.DoesNotExist:
			# No revelar si el usuario existe
			messages.success(request, 
				'Si la cuenta existe, recibirás un correo con instrucciones para recuperar tu contraseña.'
			)
			return redirect('login')
	
	return render(request, 'desarrollo/web/recuperar_contrasena.html')


def restablecer_contrasena(request, token):
	"""Vista para restablecer contraseña"""
	try:
		recuperacion = RecuperacionContrasena.objects.get(token=token)
		
		if not recuperacion.esta_vigente():
			messages.error(request, 'El enlace ha expirado. Solicita uno nuevo.')
			return redirect('recuperar_contrasena')
		
		if request.method == 'POST':
			nueva_password = request.POST.get('nueva_password', '').strip()
			confirmar_password = request.POST.get('confirmar_password', '').strip()
			
			if not nueva_password or not confirmar_password:
				messages.error(request, 'Por favor completa todos los campos.')
				return render(request, 'desarrollo/web/restablecer_contrasena.html', {'token': token})
			
			if nueva_password != confirmar_password:
				messages.error(request, 'Las contraseñas no coinciden.')
				return render(request, 'desarrollo/web/restablecer_contrasena.html', {'token': token})
			
			if len(nueva_password) < 8:
				messages.error(request, 'La contraseña debe tener al menos 8 caracteres.')
				return render(request, 'desarrollo/web/restablecer_contrasena.html', {'token': token})
			
			# Actualizar contraseña
			usuario = recuperacion.usuario
			usuario.set_password(nueva_password)
			usuario.save()
			
			# Marcar token como utilizado
			recuperacion.utilizado = True
			recuperacion.fecha_uso = timezone.now()
			recuperacion.save()
			
			messages.success(request, 'Contraseña actualizada correctamente. Por favor inicia sesión.')
			return redirect('login')
		
		return render(request, 'desarrollo/web/restablecer_contrasena.html', {'token': token})
		
	except RecuperacionContrasena.DoesNotExist:
		messages.error(request, 'Enlace inválido. Solicita uno nuevo.')
		return redirect('recuperar_contrasena')


def get_client_ip(request):
	"""Obtiene la IP del cliente"""
	x_forwarded_for = request.META.get('HTTP_X_FORWARDED_FOR')
	if x_forwarded_for:
		ip = x_forwarded_for.split(',')[0]
	else:
		ip = request.META.get('REMOTE_ADDR')
	return ip


# ============== GESTIÓN DE EMPLEADOS ==============

from .forms import (
	PersonalEmpleadosForm,
	PersonalDireccionForm,
	PersonalDepartamentoForm,
	PersonalPuestosForm,
	PersonalTipoDeContratacionForm,
	TransparenciaArchivoUploadForm,
)


@login_required(login_url='login')
def crear_empleado(request):
	"""Vista para crear un nuevo empleado"""
	if request.method == 'POST':
		form = PersonalEmpleadosForm(request.POST, request.FILES)
		if form.is_valid():
			empleado = form.save()
			
			# Si el empleado tiene usuario, password y email, enviar correo de bienvenida
			if hasattr(empleado, '_password_texto_plano') and empleado._password_texto_plano and empleado.email:
				from .email_utils import enviar_correo_bienvenida_credenciales
				if enviar_correo_bienvenida_credenciales(empleado, empleado._password_texto_plano):
					messages.success(request, f'Empleado {empleado.nombre_completo} creado exitosamente. Correo de bienvenida enviado a {empleado.email}.')
				else:
					messages.success(request, f'Empleado {empleado.nombre_completo} creado exitosamente.')
					messages.warning(request, 'No se pudo enviar el correo de bienvenida. Verifica la configuración SMTP.')
			else:
				messages.success(request, f'Empleado {empleado.nombre_completo} creado exitosamente.')
			
			return redirect('listar_empleados')
		else:
			for field, errors in form.errors.items():
				for error in errors:
					messages.error(request, f'{field}: {error}')
	else:
		form = PersonalEmpleadosForm()
	
	return render(request, 'desarrollo/empleados/crear_empleado.html', {'form': form})


@login_required(login_url='login')
def listar_empleados(request):
	"""Vista para listar todos los empleados"""
	empleados = PersonalEmpleados.objects.all().order_by('nombre_completo')
	
	# Búsqueda y filtro
	search = request.GET.get('search', '').strip()
	if search:
		empleados = empleados.filter(
			Q(nombre_completo__icontains=search) | 
			Q(usuario__icontains=search) |
			Q(email__icontains=search) |
			Q(numero_empleado__icontains=search)
		)
	
	return render(request, 'desarrollo/empleados/listar_empleados.html', {'empleados': empleados, 'search': search})


@login_required(login_url='login')
def editar_empleado(request, id_empleado):
	"""Vista para editar un empleado"""
	empleado = get_object_or_404(PersonalEmpleados, id_empleado=id_empleado)
	
	if request.method == 'POST':
		# Verificar si antes NO tenía credenciales completas
		tenia_credenciales_completas = bool(empleado.usuario and empleado.password and empleado.email)
		
		form = PersonalEmpleadosForm(request.POST, request.FILES, instance=empleado)
		if form.is_valid():
			empleado_actualizado = form.save()
			
			# Si se acaban de configurar credenciales (y antes no las tenía), enviar correo de bienvenida
			tiene_credenciales_ahora = bool(empleado_actualizado.usuario and empleado_actualizado.password and empleado_actualizado.email)
			
			if not tenia_credenciales_completas and tiene_credenciales_ahora:
				# Se configuraron credenciales por primera vez
				if hasattr(empleado_actualizado, '_password_texto_plano') and empleado_actualizado._password_texto_plano:
					from .email_utils import enviar_correo_bienvenida_credenciales
					if enviar_correo_bienvenida_credenciales(empleado_actualizado, empleado_actualizado._password_texto_plano):
						messages.success(request, f'Empleado {empleado_actualizado.nombre_completo} actualizado exitosamente. Correo de bienvenida enviado a {empleado_actualizado.email}.')
					else:
						messages.success(request, f'Empleado {empleado_actualizado.nombre_completo} actualizado exitosamente.')
						messages.warning(request, 'No se pudo enviar el correo de bienvenida. Verifica la configuración SMTP.')
				else:
					messages.success(request, f'Empleado {empleado_actualizado.nombre_completo} actualizado exitosamente.')
			else:
				messages.success(request, f'Empleado {empleado_actualizado.nombre_completo} actualizado exitosamente.')
			
			return redirect('listar_empleados')
		else:
			for field, errors in form.errors.items():
				for error in errors:
					messages.error(request, f'{field}: {error}')
	else:
		form = PersonalEmpleadosForm(instance=empleado)
	
	return render(request, 'desarrollo/empleados/editar_empleado.html', {'form': form, 'empleado': empleado})


@login_required(login_url='login')
def ver_empleado(request, id_empleado):
	"""Vista para ver detalles de un empleado"""
	empleado = get_object_or_404(PersonalEmpleados, id_empleado=id_empleado)
	return render(request, 'desarrollo/empleados/ver_empleado.html', {'empleado': empleado})


@login_required(login_url='login')
def eliminar_empleado(request, id_empleado):
	"""Vista para eliminar un empleado (cambiar a inactivo)"""
	empleado = get_object_or_404(PersonalEmpleados, id_empleado=id_empleado)
	
	if request.method == 'POST':
		empleado.activo = False
		empleado.save()
		messages.success(request, f'Empleado {empleado.nombre_completo} desactivado.')
		return redirect('listar_empleados')
	
	return render(request, 'desarrollo/empleados/confirmar_eliminar.html', {'empleado': empleado})


@login_required(login_url='login')
def reenviar_credenciales_empleado(request, id_empleado):
	"""Vista para reenviar credenciales de acceso a un empleado (AJAX)"""
	from django.http import JsonResponse
	from .email_utils import enviar_credenciales_existentes
	
	empleado = get_object_or_404(PersonalEmpleados, id_empleado=id_empleado)
	
	# Verificar que el empleado tenga al menos usuario y email
	if not empleado.usuario or not empleado.email:
		return JsonResponse({
			'success': False, 
			'message': 'El empleado no tiene usuario y/o email configurados.'
		}, status=400)
	
	# Enviar correo con nueva contraseña temporal
	success, password_temporal = enviar_credenciales_existentes(empleado)
	
	if success:
		return JsonResponse({
			'success': True, 
			'message': f'Contraseña temporal generada y enviada a {empleado.email}'
		})
	else:
		return JsonResponse({
			'success': False, 
			'message': 'Hubo un error al enviar el correo. Verifica la configuración SMTP.'
		}, status=500)


# ============== RECURSOS HUMANOS: DIRECCIONES Y DEPARTAMENTOS ==============

@login_required(login_url='login')
def listar_direcciones(request):
	from .models import PersonalDireccion

	search = request.GET.get('search', '').strip()
	direcciones = PersonalDireccion.objects.all().order_by('direccion')

	if search:
		direcciones = direcciones.filter(
			Q(direccion__icontains=search) |
			Q(descripcion__icontains=search)
		)

	return render(request, 'desarrollo/recursos_humanos/listar_direcciones.html', {
		'direcciones': direcciones,
		'search': search,
	})


@login_required(login_url='login')
def crear_direccion(request):
	if request.method == 'POST':
		form = PersonalDireccionForm(request.POST)
		if form.is_valid():
			direccion = form.save(commit=False)
			usuario_actual = getattr(request.user, 'usuario', None) or getattr(request.user, 'username', 'sistema')
			direccion.usuario_captura = direccion.usuario_captura or usuario_actual
			direccion.usuario_modificacion = usuario_actual
			direccion.save()
			messages.success(request, 'Dirección creada exitosamente.')
			return redirect('listar_direcciones')
	else:
		form = PersonalDireccionForm()

	return render(request, 'desarrollo/recursos_humanos/form_direccion.html', {
		'form': form,
		'titulo': 'Nueva Dirección',
		'accion': 'Crear'
	})


@login_required(login_url='login')
def editar_direccion(request, iddireccion):
	from .models import PersonalDireccion
	direccion = get_object_or_404(PersonalDireccion, iddireccion=iddireccion)

	if request.method == 'POST':
		form = PersonalDireccionForm(request.POST, instance=direccion)
		if form.is_valid():
			direccion = form.save(commit=False)
			usuario_actual = getattr(request.user, 'usuario', None) or getattr(request.user, 'username', 'sistema')
			direccion.usuario_modificacion = usuario_actual
			direccion.save()
			messages.success(request, 'Dirección actualizada exitosamente.')
			return redirect('listar_direcciones')
	else:
		form = PersonalDireccionForm(instance=direccion)

	return render(request, 'desarrollo/recursos_humanos/form_direccion.html', {
		'form': form,
		'direccion': direccion,
		'titulo': 'Editar Dirección',
		'accion': 'Guardar cambios'
	})


@login_required(login_url='login')
def cambiar_estado_direccion(request, iddireccion):
	from .models import PersonalDireccion
	direccion = get_object_or_404(PersonalDireccion, iddireccion=iddireccion)

	if request.method == 'POST':
		direccion.activo = not direccion.activo
		usuario_actual = getattr(request.user, 'usuario', None) or getattr(request.user, 'username', 'sistema')
		direccion.usuario_modificacion = usuario_actual
		direccion.save()
		estado = 'activada' if direccion.activo else 'desactivada'
		messages.success(request, f'Dirección {estado} exitosamente.')

	return redirect('listar_direcciones')


@login_required(login_url='login')
def listar_departamentos(request):
	from .models import PersonalDepartamento

	search = request.GET.get('search', '').strip()
	departamentos = PersonalDepartamento.objects.select_related('iddireccion').all().order_by('iddireccion__direccion', 'departamento')

	if search:
		departamentos = departamentos.filter(
			Q(departamento__icontains=search) |
			Q(descripcion__icontains=search) |
			Q(iddireccion__direccion__icontains=search)
		)

	return render(request, 'desarrollo/recursos_humanos/listar_departamentos.html', {
		'departamentos': departamentos,
		'search': search,
	})


@login_required(login_url='login')
def crear_departamento(request):
	if request.method == 'POST':
		form = PersonalDepartamentoForm(request.POST)
		if form.is_valid():
			departamento = form.save(commit=False)
			usuario_actual = getattr(request.user, 'usuario', None) or getattr(request.user, 'username', 'sistema')
			departamento.usuario_captura = departamento.usuario_captura or usuario_actual
			departamento.usuario_modificacion = usuario_actual
			departamento.save()
			messages.success(request, 'Departamento creado exitosamente.')
			return redirect('listar_departamentos')
	else:
		form = PersonalDepartamentoForm()

	return render(request, 'desarrollo/recursos_humanos/form_departamento.html', {
		'form': form,
		'titulo': 'Nuevo Departamento',
		'accion': 'Crear'
	})


@login_required(login_url='login')
def editar_departamento(request, iddepartamento):
	from .models import PersonalDepartamento
	departamento = get_object_or_404(PersonalDepartamento, iddepartamento=iddepartamento)

	if request.method == 'POST':
		form = PersonalDepartamentoForm(request.POST, instance=departamento)
		if form.is_valid():
			departamento = form.save(commit=False)
			usuario_actual = getattr(request.user, 'usuario', None) or getattr(request.user, 'username', 'sistema')
			departamento.usuario_modificacion = usuario_actual
			departamento.save()
			messages.success(request, 'Departamento actualizado exitosamente.')
			return redirect('listar_departamentos')
	else:
		form = PersonalDepartamentoForm(instance=departamento)

	return render(request, 'desarrollo/recursos_humanos/form_departamento.html', {
		'form': form,
		'departamento': departamento,
		'titulo': 'Editar Departamento',
		'accion': 'Guardar cambios'
	})


@login_required(login_url='login')
def cambiar_estado_departamento(request, iddepartamento):
	from .models import PersonalDepartamento
	departamento = get_object_or_404(PersonalDepartamento, iddepartamento=iddepartamento)

	if request.method == 'POST':
		departamento.activo = not departamento.activo
		usuario_actual = getattr(request.user, 'usuario', None) or getattr(request.user, 'username', 'sistema')
		departamento.usuario_modificacion = usuario_actual
		departamento.save()
		estado = 'activado' if departamento.activo else 'desactivado'
		messages.success(request, f'Departamento {estado} exitosamente.')

	return redirect('listar_departamentos')


@login_required(login_url='login')
def listar_puestos(request):
	from .models import PersonalPuestos

	search = request.GET.get('search', '').strip()
	puestos = PersonalPuestos.objects.all().order_by('nombre')

	if search:
		puestos = puestos.filter(
			Q(nombre__icontains=search) |
			Q(descripcion__icontains=search)
		)

	return render(request, 'desarrollo/recursos_humanos/listar_puestos.html', {
		'puestos': puestos,
		'search': search,
	})


@login_required(login_url='login')
def crear_puesto(request):
	if request.method == 'POST':
		form = PersonalPuestosForm(request.POST)
		if form.is_valid():
			puesto = form.save(commit=False)
			usuario_actual = getattr(request.user, 'usuario', None) or getattr(request.user, 'username', 'sistema')
			puesto.usuario_captura = puesto.usuario_captura or usuario_actual
			puesto.usuario_modificacion = usuario_actual
			puesto.save()
			messages.success(request, 'Puesto creado exitosamente.')
			return redirect('listar_puestos')
	else:
		form = PersonalPuestosForm()

	return render(request, 'desarrollo/recursos_humanos/form_puesto.html', {
		'form': form,
		'titulo': 'Nuevo Puesto',
		'accion': 'Crear'
	})


@login_required(login_url='login')
def editar_puesto(request, idpuesto):
	from .models import PersonalPuestos
	puesto = get_object_or_404(PersonalPuestos, idpuesto=idpuesto)

	if request.method == 'POST':
		form = PersonalPuestosForm(request.POST, instance=puesto)
		if form.is_valid():
			puesto = form.save(commit=False)
			usuario_actual = getattr(request.user, 'usuario', None) or getattr(request.user, 'username', 'sistema')
			puesto.usuario_modificacion = usuario_actual
			puesto.save()
			messages.success(request, 'Puesto actualizado exitosamente.')
			return redirect('listar_puestos')
	else:
		form = PersonalPuestosForm(instance=puesto)

	return render(request, 'desarrollo/recursos_humanos/form_puesto.html', {
		'form': form,
		'puesto': puesto,
		'titulo': 'Editar Puesto',
		'accion': 'Guardar cambios'
	})


@login_required(login_url='login')
def cambiar_estado_puesto(request, idpuesto):
	from .models import PersonalPuestos
	puesto = get_object_or_404(PersonalPuestos, idpuesto=idpuesto)

	if request.method == 'POST':
		puesto.activo = not puesto.activo
		usuario_actual = getattr(request.user, 'usuario', None) or getattr(request.user, 'username', 'sistema')
		puesto.usuario_modificacion = usuario_actual
		puesto.save()
		estado = 'activado' if puesto.activo else 'desactivado'
		messages.success(request, f'Puesto {estado} exitosamente.')

	return redirect('listar_puestos')


@login_required(login_url='login')
def listar_tipos_contratacion(request):
	from .models import PersonalTipoDeContratacion

	search = request.GET.get('search', '').strip()
	tipos = PersonalTipoDeContratacion.objects.all().order_by('nombre')

	if search:
		tipos = tipos.filter(
			Q(nombre__icontains=search) |
			Q(descripcion__icontains=search)
		)

	return render(request, 'desarrollo/recursos_humanos/listar_tipos_contratacion.html', {
		'tipos': tipos,
		'search': search,
	})


@login_required(login_url='login')
def crear_tipo_contratacion(request):
	if request.method == 'POST':
		form = PersonalTipoDeContratacionForm(request.POST)
		if form.is_valid():
			tipo = form.save(commit=False)
			usuario_actual = getattr(request.user, 'usuario', None) or getattr(request.user, 'username', 'sistema')
			tipo.usuario_captura = tipo.usuario_captura or usuario_actual
			tipo.usuario_modificacion = usuario_actual
			tipo.save()
			messages.success(request, 'Tipo de contratación creado exitosamente.')
			return redirect('listar_tipos_contratacion')
	else:
		form = PersonalTipoDeContratacionForm()

	return render(request, 'desarrollo/recursos_humanos/form_tipo_contratacion.html', {
		'form': form,
		'titulo': 'Nuevo Tipo de Contratación',
		'accion': 'Crear'
	})


@login_required(login_url='login')
def editar_tipo_contratacion(request, idtipodecontratacion):
	from .models import PersonalTipoDeContratacion
	tipo = get_object_or_404(PersonalTipoDeContratacion, idtipodecontratacion=idtipodecontratacion)

	if request.method == 'POST':
		form = PersonalTipoDeContratacionForm(request.POST, instance=tipo)
		if form.is_valid():
			tipo = form.save(commit=False)
			usuario_actual = getattr(request.user, 'usuario', None) or getattr(request.user, 'username', 'sistema')
			tipo.usuario_modificacion = usuario_actual
			tipo.save()
			messages.success(request, 'Tipo de contratación actualizado exitosamente.')
			return redirect('listar_tipos_contratacion')
	else:
		form = PersonalTipoDeContratacionForm(instance=tipo)

	return render(request, 'desarrollo/recursos_humanos/form_tipo_contratacion.html', {
		'form': form,
		'tipo': tipo,
		'titulo': 'Editar Tipo de Contratación',
		'accion': 'Guardar cambios'
	})


@login_required(login_url='login')
def cambiar_estado_tipo_contratacion(request, idtipodecontratacion):
	from .models import PersonalTipoDeContratacion
	tipo = get_object_or_404(PersonalTipoDeContratacion, idtipodecontratacion=idtipodecontratacion)

	if request.method == 'POST':
		tipo.activo = not tipo.activo
		usuario_actual = getattr(request.user, 'usuario', None) or getattr(request.user, 'username', 'sistema')
		tipo.usuario_modificacion = usuario_actual
		tipo.save()
		estado = 'activado' if tipo.activo else 'desactivado'
		messages.success(request, f'Tipo de contratación {estado} exitosamente.')

	return redirect('listar_tipos_contratacion')


@login_required(login_url='login')
def get_departamentos_por_direccion(request):
	"""Vista AJAX que devuelve los departamentos filtrados por dirección"""
	from django.http import JsonResponse
	from .models import PersonalDepartamento
	
	iddireccion = request.GET.get('iddireccion')
	
	if not iddireccion:
		return JsonResponse({'departamentos': []})
	
	try:
		departamentos = PersonalDepartamento.objects.filter(
			iddireccion_id=iddireccion,
			activo=True
		).values('iddepartamento', 'departamento').order_by('departamento')
		
		return JsonResponse({
			'departamentos': list(departamentos)
		})
	except Exception as e:
		return JsonResponse({'error': str(e)}, status=400)

# ===================== VISTAS DE SEGURIDAD DEL SISTEMA =====================

@login_required(login_url='login')
def listar_usuarios_sistema(request):
	"""Lista los usuarios del sistema con filtros"""
	from django.db.models import Q
	from .forms import FiltroUsuariosDelSistemaForm
	
	# Verificar permisos (solo admin)
	if not _usuario_es_admin_sistema(request.user):
		messages.error(request, 'No tienes permisos para acceder a Usuarios del Sistema.')
		return redirect('dashboard')
	
	# Obtener parámetros de filtro
	form = FiltroUsuariosDelSistemaForm(request.GET or None)
	
	# Usuarios activos: empleados con usuario, email y password configurado
	usuarios = PersonalEmpleados.objects.filter(
		activo=True,
		usuario__isnull=False,
		email__isnull=False
	).exclude(
		usuario='',
		email=''
	).order_by('-fecha_ultimo_login', 'nombre_completo')
	
	# Aplicar filtros de búsqueda
	if request.GET:
		busqueda = request.GET.get('busqueda', '').strip()
		if busqueda:
			usuarios = usuarios.filter(
				Q(usuario__icontains=busqueda) |
				Q(email__icontains=busqueda) |
				Q(nombre_completo__icontains=busqueda)
			)
		
		# Filtro de estado
		estado = request.GET.get('estado', '').strip()
		if estado == 'activo':
			usuarios = usuarios.filter(is_active=True)
		elif estado == 'inactivo':
			usuarios = usuarios.filter(is_active=False)
	
	# Empleados candidatos: sin usuario, email o ambos vacíos
	candidatos = PersonalEmpleados.objects.filter(
		activo=True
	).filter(
		Q(usuario__isnull=True) | Q(usuario='') |
		Q(email__isnull=True) | Q(email='')
	).order_by('nombre_completo')
	
	# Aplicar búsqueda a candidatos si hay filtro
	if request.GET:
		busqueda = request.GET.get('busqueda', '').strip()
		if busqueda:
			candidatos = candidatos.filter(
				Q(nombre_completo__icontains=busqueda) |
				Q(numero_empleado__icontains=busqueda)
			)
	
	context = {
		'usuarios': usuarios,
		'candidatos': candidatos,
		'form': form,
		'titulo': 'Usuarios del Sistema',
		'total_usuarios': usuarios.count(),
		'total_candidatos': candidatos.count(),
	}
	
	return render(request, 'desarrollo/seguridad/listar_usuarios.html', context)


@login_required(login_url='login')
def crear_usuario_sistema(request):
	"""Crear un nuevo usuario del sistema"""
	from django.core.mail import send_mail
	from django.conf import settings
	from .models import UsuariosDelSistema, ConfiguracionSistema
	from .forms import UsuariosDelSistemaForm
	
	# Verificar permisos (solo admin)
	if not _usuario_es_admin_sistema(request.user):
		messages.error(request, 'No tienes permisos para crear usuarios del sistema.')
		return redirect('dashboard')
	
	if request.method == 'POST':
		form = UsuariosDelSistemaForm(request.POST)
		if form.is_valid():
			usuario = form.save(commit=False)
			usuario.creado_por = request.user
			usuario.save()
			
			# Obtener configuración del sistema
			try:
				config = ConfiguracionSistema.objects.first()
				nombre_empresa = config.nombre_corto if config else 'ITAVU'
				email_desde = config.email_desde if config else settings.DEFAULT_FROM_EMAIL
			except:
				nombre_empresa = 'ITAVU'
				email_desde = settings.DEFAULT_FROM_EMAIL
			
			# Enviar correo de bienvenida
			try:
				asunto = f'Bienvenido al Sistema de {nombre_empresa}'
				mensaje = f"""
				¡Hola {usuario.id_empleado.nombre}!
				
				Te damos la bienvenida al equipo de {nombre_empresa}.
				
				Tu acceso al sistema ha sido creado exitosamente.
				
				Datos de acceso:
				Usuario: {usuario.usuario}
				Correo: {usuario.correo}
				
				Por favor, cambia tu contraseña en el primer inicio de sesión.
				
				¡Bienvenido!
				"""
				
				send_mail(
					asunto,
					mensaje,
					email_desde,
					[usuario.correo],
					fail_silently=False,
				)
			except Exception as e:
				print(f"Error al enviar correo: {e}")
			
			messages.success(request, f'Usuario {usuario.usuario} creado exitosamente. Correo de bienvenida enviado.')
			return redirect('listar_usuarios_sistema')
	else:
		form = UsuariosDelSistemaForm()
	
	context = {
		'form': form,
		'titulo': 'Crear Usuario del Sistema',
		'modo': 'crear',
	}
	
	return render(request, 'desarrollo/seguridad/formulario_usuario.html', context)


@login_required(login_url='login')
def editar_usuario_sistema(request, id_usuario):
	"""Editar usuario del sistema"""
	from django.shortcuts import get_object_or_404
	from .models import UsuariosDelSistema
	from .forms import UsuariosDelSistemaForm
	
	# Verificar permisos (solo admin)
	if not _usuario_es_admin_sistema(request.user):
		messages.error(request, 'No tienes permisos para editar usuarios del sistema.')
		return redirect('dashboard')
	
	usuario = get_object_or_404(UsuariosDelSistema, id_usuario_sistema=id_usuario)
	
	if request.method == 'POST':
		form = UsuariosDelSistemaForm(request.POST, instance=usuario)
		if form.is_valid():
			usuario_actualizado = form.save()
			messages.success(request, f'Usuario {usuario_actualizado.usuario} actualizado exitosamente.')
			return redirect('listar_usuarios_sistema')
	else:
		form = UsuariosDelSistemaForm(instance=usuario)
	
	context = {
		'form': form,
		'titulo': 'Editar Usuario del Sistema',
		'modo': 'editar',
		'usuario': usuario,
	}
	
	return render(request, 'desarrollo/seguridad/formulario_usuario.html', context)


@login_required(login_url='login')
def configuracion_sistema(request):
	"""Configuración del sistema (SMTP, datos de empresa, etc.)"""
	from .models import ConfiguracionSistema, UsuariosDelSistema
	from .forms import ConfiguracionSistemaForm
	from django.shortcuts import get_object_or_404
	
	# Verificar permisos (solo admin)
	if not _usuario_es_admin_sistema(request.user):
		messages.error(request, 'No tienes permisos para acceder a Configuración del Sistema.')
		return redirect('dashboard')
	
	# Obtener o crear configuración
	config, created = ConfiguracionSistema.objects.get_or_create(pk=1)

	def _aplicar_departamento_soporte_desde_post(configuracion, post_data):
		"""Aplica y guarda el departamento de soporte desde POST de forma independiente."""
		if 'departamento_soporte_mantenimiento' not in post_data:
			return False

		dept_id = (post_data.get('departamento_soporte_mantenimiento') or '').strip()

		if dept_id == '':
			if configuracion.departamento_soporte_mantenimiento_id is not None:
				configuracion.departamento_soporte_mantenimiento = None
				configuracion.save(update_fields=['departamento_soporte_mantenimiento', 'fecha_modificacion'])
				return True
			return False

		try:
			departamento = PersonalDepartamento.objects.get(iddepartamento=dept_id, activo=True)
		except PersonalDepartamento.DoesNotExist:
			return False

		if configuracion.departamento_soporte_mantenimiento_id != departamento.iddepartamento:
			configuracion.departamento_soporte_mantenimiento = departamento
			configuracion.save(update_fields=['departamento_soporte_mantenimiento', 'fecha_modificacion'])
			return True

		return False
	
	if request.method == 'POST':
		cambios_departamento_soporte = _aplicar_departamento_soporte_desde_post(config, request.POST)
		# Asegurar que el form use la instancia más reciente (incluyendo guardado parcial)
		if cambios_departamento_soporte:
			config.refresh_from_db()

		form = ConfiguracionSistemaForm(request.POST, request.FILES, instance=config)
		if form.is_valid():
			form.save()
			messages.success(request, 'Configuración del sistema actualizada exitosamente.')
			return redirect('configuracion_sistema')
		else:
			if cambios_departamento_soporte:
				messages.warning(
					request,
					'Se guardó el Departamento de Soporte para Mantenimiento, pero hay otros campos con errores por corregir.'
				)
	else:
		form = ConfiguracionSistemaForm(instance=config)

	total_usuarios = UsuariosDelSistema.objects.count()
	usuarios_activos = UsuariosDelSistema.objects.filter(activo=True).count()
	usuarios_inactivos = total_usuarios - usuarios_activos
	
	context = {
		'form': form,
		'titulo': 'Configuración del Sistema',
		'config': config,
		'total_usuarios': total_usuarios,
		'usuarios_activos': usuarios_activos,
		'usuarios_inactivos': usuarios_inactivos,
	}
	
	return render(request, 'desarrollo/seguridad/configuracion.html', context)


@login_required(login_url='login')
def probar_smtp(request):
	"""Prueba la configuración SMTP enviando un correo de prueba"""
	from django.http import JsonResponse
	from django.core.mail import send_mail
	from django.conf import settings
	from .models import ConfiguracionSistema
	
	# Verificar permisos (solo admin)
	if not _usuario_es_admin_sistema(request.user):
		return JsonResponse({'success': False, 'message': 'No tienes permisos para esta acción.'}, status=403)
	
	if request.method != 'POST':
		return JsonResponse({'success': False, 'message': 'Método no permitido.'}, status=405)
	
	email_prueba = request.POST.get('email_prueba', '').strip()
	
	if not email_prueba:
		return JsonResponse({'success': False, 'message': 'Por favor proporciona un correo electrónico.'})
	
	try:
		# Obtener configuración
		config = ConfiguracionSistema.objects.first()
		
		if not config:
			return JsonResponse({
				'success': False, 
				'message': 'No se ha configurado el sistema. Por favor guarda la configuración primero.'
			})
		
		if not config.smtp_host or not config.email_desde:
			return JsonResponse({
				'success': False, 
				'message': 'Configuración SMTP incompleta. Por favor completa todos los campos requeridos.'
			})
		
		# Configurar SMTP temporalmente
		from django.core.mail import get_connection
		
		connection = get_connection(
			backend='django.core.mail.backends.smtp.EmailBackend',
			host=config.smtp_host,
			port=config.smtp_port,
			username=config.smtp_usuario,
			password=config.smtp_contrasena,
			use_tls=config.smtp_usar_tls,
			fail_silently=False,
		)
		
		# Enviar correo de prueba
		asunto = f'Prueba de Configuración SMTP - {config.nombre_corto}'
		mensaje = f'''
¡Hola!

Este es un correo de prueba del sistema {config.nombre_corto}.

Si recibes este mensaje, significa que la configuración SMTP está funcionando correctamente.

Detalles de la configuración:
- Servidor SMTP: {config.smtp_host}
- Puerto: {config.smtp_port}
- Usuario: {config.smtp_usuario}
- TLS: {'Sí' if config.smtp_usar_tls else 'No'}

Saludos,
Sistema {config.nombre_corto}
		'''
		
		send_mail(
			asunto,
			mensaje,
			config.email_desde,
			[email_prueba],
			connection=connection,
			fail_silently=False,
		)
		
		return JsonResponse({
			'success': True, 
			'message': f'¡Correo de prueba enviado exitosamente a {email_prueba}! Revisa tu bandeja de entrada.'
		})
		
	except Exception as e:
		return JsonResponse({
			'success': False, 
			'message': f'Error al enviar correo: {str(e)}'
		})


@login_required(login_url='login')
def monitor_cola_correos(request):
	"""Vista para monitorear la cola de correos (admin)"""
	from .models import ColaCorreos
	from django.db.models import Count, Q
	from datetime import datetime, timedelta
	
	# Verificar permisos (solo admin)
	if not _usuario_es_admin_sistema(request.user):
		messages.error(request, 'No tienes permisos para acceder a esta sección.')
		return redirect('dashboard')
	
	# Estadísticas generales
	total_correos = ColaCorreos.objects.count()
	correos_pendientes = ColaCorreos.objects.filter(estado='pendiente').count()
	correos_enviados = ColaCorreos.objects.filter(estado='enviado').count()
	correos_error = ColaCorreos.objects.filter(estado='error').count()
	
	# Correos enviados hoy
	inicio_dia = timezone.now().replace(hour=0, minute=0, second=0, microsecond=0)
	correos_hoy = ColaCorreos.objects.filter(
		estado='enviado',
		fecha_envio__gte=inicio_dia
	).count()
	
	# Últimos correos
	ultimos_correos = ColaCorreos.objects.all().order_by('-fecha_creacion')[:20]
	porcentaje_hoy = min(round((correos_hoy / 2000) * 100, 2), 100)
	
	context = {
		'total_correos': total_correos,
		'correos_pendientes': correos_pendientes,
		'correos_enviados': correos_enviados,
		'correos_error': correos_error,
		'correos_hoy': correos_hoy,
		'limite_diario': 2000,
		'porcentaje_hoy': porcentaje_hoy,
		'ultimos_correos': ultimos_correos,
	}
	
	return render(request, 'desarrollo/seguridad/monitor_cola_correos.html', context)


@login_required(login_url='login')
def procesar_cola_ahora(request):
	"""Vista AJAX para procesar la cola de correos manualmente"""
	from django.http import JsonResponse
	from .email_utils import procesar_cola_correos
	
	# Verificar permisos (solo admin)
	if not _usuario_es_admin_sistema(request.user):
		return JsonResponse({'success': False, 'message': 'No tienes permisos'}, status=403)
	
	if request.method != 'POST':
		return JsonResponse({'success': False, 'message': 'Método no permitido'}, status=405)
	
	resultado = procesar_cola_correos(limite_diario=2000)
	
	return JsonResponse({
		'success': True,
		'enviados': resultado['enviados'],
		'errores': resultado['errores'],
		'pendientes': resultado['pendientes'],
		'message': f"Procesamiento completado. Enviados: {resultado['enviados']}, Errores: {resultado['errores']}"
	})


@login_required(login_url='login')
def api_estadisticas_cola(request):
	"""API para obtener estadísticas de la cola de correos en JSON"""
	from .models import ColaCorreos
	
	# Verificar permisos (solo admin)
	if not _usuario_es_admin_sistema(request.user):
		return JsonResponse({'success': False, 'message': 'No tienes permisos'}, status=403)
	
	# Estadísticas generales
	total_correos = ColaCorreos.objects.count()
	correos_pendientes = ColaCorreos.objects.filter(estado='pendiente').count()
	correos_enviados = ColaCorreos.objects.filter(estado='enviado').count()
	correos_error = ColaCorreos.objects.filter(estado='error').count()
	
	# Correos enviados hoy
	inicio_dia = timezone.now().replace(hour=0, minute=0, second=0, microsecond=0)
	correos_hoy = ColaCorreos.objects.filter(
		estado='enviado',
		fecha_envio__gte=inicio_dia
	).count()
	
	# Por tipo de correo
	por_tipo = {}
	for tipo, nombre in ColaCorreos.TIPO_CORREO_CHOICES:
		por_tipo[nombre] = ColaCorreos.objects.filter(tipo_correo=tipo).count()
	
	return JsonResponse({
		'success': True,
		'total': total_correos,
		'pendientes': correos_pendientes,
		'enviados': correos_enviados,
		'errores': correos_error,
		'enviados_hoy': correos_hoy,
		'limite_diario': 2000,
		'porcentaje_diario': round((correos_hoy / 2000) * 100, 2),
		'por_tipo': por_tipo,
	})


# ==================== VISTAS DE PATRIMONIO ====================

@login_required(login_url='login')
def listar_bienes(request):
	"""Lista todos los bienes del Instituto"""
	from .models import PatrimonioBienesDelInstituto

	try:
		bienes = PatrimonioBienesDelInstituto.objects.all().order_by('-fecha_creacion')
		# Validación temprana para detectar tabla inexistente
		bienes.count()
	except (ProgrammingError, OperationalError):
		messages.warning(
			request,
			'El módulo de Patrimonio aún no está migrado en la base de datos. Aplica las migraciones para habilitar esta sección.'
		)
		bienes = PatrimonioBienesDelInstituto.objects.none()
	
	# Búsqueda y filtrado
	busqueda = request.GET.get('busqueda', '')
	estado_filtro = request.GET.get('estado', '')
	
	if busqueda:
		bienes = bienes.filter(
			Q(numero_inventario_itavu__icontains=busqueda) |
			Q(numero_inventario_gobierno__icontains=busqueda) |
			Q(descripcion__icontains=busqueda) |
			Q(serie__icontains=busqueda) |
			Q(numero_factura__icontains=busqueda)
		)
	
	if estado_filtro:
		bienes = bienes.filter(activo=(estado_filtro == 'activo'))

	total_bienes = bienes.count()
	
	context = {
		'bienes': bienes,
		'total_bienes': total_bienes,
		'busqueda': busqueda,
		'estado_filtro': estado_filtro,
	}
	return render(request, 'desarrollo/patrimonio/listar_bienes.html', context)


@login_required(login_url='login')
def crear_bien(request):
	"""Crear un nuevo bien del Instituto"""
	from .forms import PatrimonioBienesDelInstitutoForm
	from .models import PatrimonioBienesDelInstituto
	
	if request.method == 'POST':
		form = PatrimonioBienesDelInstitutoForm(request.POST, request.FILES)
		if form.is_valid():
			bien = form.save(commit=False)
			bien.usuario_captura = request.user.usuario if hasattr(request.user, 'usuario') else str(request.user)
			bien.save()
			messages.success(request, f'Bien "{bien.numero_inventario_itavu}" creado exitosamente.')
			return redirect('listar_bienes')
	else:
		form = PatrimonioBienesDelInstitutoForm()
	
	context = {'form': form, 'titulo': 'Crear Bien del Instituto'}
	return render(request, 'desarrollo/patrimonio/form_bien.html', context)


@login_required(login_url='login')
def editar_bien(request, idbien):
	"""Editar un bien del Instituto"""
	from .forms import PatrimonioBienesDelInstitutoForm
	from .models import PatrimonioBienesDelInstituto
	
	bien = get_object_or_404(PatrimonioBienesDelInstituto, idbien=idbien)
	
	if request.method == 'POST':
		form = PatrimonioBienesDelInstitutoForm(request.POST, request.FILES, instance=bien)
		if form.is_valid():
			bien = form.save(commit=False)
			bien.usuario_modificacion = request.user.usuario if hasattr(request.user, 'usuario') else str(request.user)
			bien.save()
			messages.success(request, f'Bien "{bien.numero_inventario_itavu}" actualizado exitosamente.')
			return redirect('listar_bienes')
	else:
		form = PatrimonioBienesDelInstitutoForm(instance=bien)
	
	context = {'form': form, 'bien': bien, 'titulo': 'Editar Bien del Instituto'}
	return render(request, 'desarrollo/patrimonio/form_bien.html', context)


@login_required(login_url='login')
def cambiar_estado_bien(request, idbien):
	"""Cambiar estado (activo/inactivo) de un bien"""
	from .models import PatrimonioBienesDelInstituto
	
	bien = get_object_or_404(PatrimonioBienesDelInstituto, idbien=idbien)
	bien.activo = not bien.activo
	bien.usuario_modificacion = request.user.usuario if hasattr(request.user, 'usuario') else str(request.user)
	bien.save()
	
	estado_texto = 'activado' if bien.activo else 'desactivado'
	messages.success(request, f'Bien "{bien.numero_inventario_itavu}" {estado_texto} exitosamente.')
	return redirect('listar_bienes')


# ============== CATÁLOGOS DE PATRIMONIO: MARCAS ==============

@login_required(login_url='login')
def listar_marcas(request):
	from .models import CatalogosMarcas
	from .forms import CatalogosMarcasForm
	
	search = request.GET.get('search', '').strip()
	marcas = CatalogosMarcas.objects.all().order_by('nombre')
	
	if search:
		marcas = marcas.filter(Q(nombre__icontains=search) | Q(descripcion__icontains=search))
	
	return render(request, 'desarrollo/patrimonio/listar_marcas.html', {'marcas': marcas, 'search': search})


@login_required(login_url='login')
def crear_marca(request):
	from .forms import CatalogosMarcasForm
	
	if request.method == 'POST':
		form = CatalogosMarcasForm(request.POST)
		if form.is_valid():
			form.save()
			messages.success(request, 'Marca creada exitosamente.')
			return redirect('listar_marcas')
	else:
		form = CatalogosMarcasForm()
	
	return render(request, 'desarrollo/patrimonio/form_marca.html', {'form': form, 'titulo': 'Nueva Marca', 'accion': 'Crear'})


@login_required(login_url='login')
def editar_marca(request, idmarca):
	from .models import CatalogosMarcas
	from .forms import CatalogosMarcasForm
	
	marca = get_object_or_404(CatalogosMarcas, idmarca=idmarca)
	
	if request.method == 'POST':
		form = CatalogosMarcasForm(request.POST, instance=marca)
		if form.is_valid():
			form.save()
			messages.success(request, 'Marca actualizada exitosamente.')
			return redirect('listar_marcas')
	else:
		form = CatalogosMarcasForm(instance=marca)
	
	return render(request, 'desarrollo/patrimonio/form_marca.html', {'form': form, 'marca': marca, 'titulo': 'Editar Marca', 'accion': 'Guardar cambios'})


@login_required(login_url='login')
def cambiar_estado_marca(request, idmarca):
	from .models import CatalogosMarcas
	
	marca = get_object_or_404(CatalogosMarcas, idmarca=idmarca)
	
	if request.method == 'POST':
		marca.activo = not marca.activo
		marca.save()
		estado = 'activada' if marca.activo else 'desactivada'
		messages.success(request, f'Marca {estado} exitosamente.')
	
	return redirect('listar_marcas')


# ============== CATÁLOGOS DE PATRIMONIO: PROVEEDORES ==============

@login_required(login_url='login')
def listar_proveedores(request):
	from .models import PatrimonioProveedor
	
	search = request.GET.get('search', '').strip()
	proveedores = PatrimonioProveedor.objects.all().order_by('nombre')
	
	if search:
		proveedores = proveedores.filter(Q(nombre__icontains=search) | Q(rfc__icontains=search) | Q(descripcion__icontains=search))
	
	return render(request, 'desarrollo/patrimonio/listar_proveedores.html', {'proveedores': proveedores, 'search': search})


@login_required(login_url='login')
def crear_proveedor(request):
	from .forms import PatrimonioProveedorForm
	
	if request.method == 'POST':
		form = PatrimonioProveedorForm(request.POST)
		if form.is_valid():
			form.save()
			messages.success(request, 'Proveedor creado exitosamente.')
			return redirect('listar_proveedores')
	else:
		form = PatrimonioProveedorForm()
	
	return render(request, 'desarrollo/patrimonio/form_proveedor.html', {'form': form, 'titulo': 'Nuevo Proveedor', 'accion': 'Crear'})


@login_required(login_url='login')
def editar_proveedor(request, idproveedor):
	from .models import PatrimonioProveedor
	from .forms import PatrimonioProveedorForm
	
	proveedor = get_object_or_404(PatrimonioProveedor, idproveedor=idproveedor)
	
	if request.method == 'POST':
		form = PatrimonioProveedorForm(request.POST, instance=proveedor)
		if form.is_valid():
			form.save()
			messages.success(request, 'Proveedor actualizado exitosamente.')
			return redirect('listar_proveedores')
	else:
		form = PatrimonioProveedorForm(instance=proveedor)
	
	return render(request, 'desarrollo/patrimonio/form_proveedor.html', {'form': form, 'proveedor': proveedor, 'titulo': 'Editar Proveedor', 'accion': 'Guardar cambios'})


@login_required(login_url='login')
def cambiar_estado_proveedor(request, idproveedor):
	from .models import PatrimonioProveedor
	
	proveedor = get_object_or_404(PatrimonioProveedor, idproveedor=idproveedor)
	
	if request.method == 'POST':
		proveedor.activo = not proveedor.activo
		proveedor.save()
		estado = 'activado' if proveedor.activo else 'desactivado'
		messages.success(request, f'Proveedor {estado} exitosamente.')
	
	return redirect('listar_proveedores')


# ============== CATÁLOGOS DE PATRIMONIO: CLASIFICACIÓN SERAP ==============

@login_required(login_url='login')
def listar_clasificaciones_serap(request):
	from .models import PatrimonioClasificacionSerap
	
	search = request.GET.get('search', '').strip()
	clasificaciones = PatrimonioClasificacionSerap.objects.all().order_by('nombre')
	
	if search:
		clasificaciones = clasificaciones.filter(Q(nombre__icontains=search) | Q(descripcion__icontains=search))
	
	return render(request, 'desarrollo/patrimonio/listar_clasificaciones_serap.html', {'clasificaciones': clasificaciones, 'search': search})


@login_required(login_url='login')
def crear_clasificacion_serap(request):
	from .forms import PatrimonioClasificacionSerapForm
	
	if request.method == 'POST':
		form = PatrimonioClasificacionSerapForm(request.POST)
		if form.is_valid():
			form.save()
			messages.success(request, 'Clasificación SERAP creada exitosamente.')
			return redirect('listar_clasificaciones_serap')
	else:
		form = PatrimonioClasificacionSerapForm()
	
	return render(request, 'desarrollo/patrimonio/form_clasificacion_serap.html', {'form': form, 'titulo': 'Nueva Clasificación SERAP', 'accion': 'Crear'})


@login_required(login_url='login')
def editar_clasificacion_serap(request, idclasificacion_serap):
	from .models import PatrimonioClasificacionSerap
	from .forms import PatrimonioClasificacionSerapForm
	
	clasificacion = get_object_or_404(PatrimonioClasificacionSerap, idclasificacion_serap=idclasificacion_serap)
	
	if request.method == 'POST':
		form = PatrimonioClasificacionSerapForm(request.POST, instance=clasificacion)
		if form.is_valid():
			form.save()
			messages.success(request, 'Clasificación SERAP actualizada exitosamente.')
			return redirect('listar_clasificaciones_serap')
	else:
		form = PatrimonioClasificacionSerapForm(instance=clasificacion)
	
	return render(request, 'desarrollo/patrimonio/form_clasificacion_serap.html', {'form': form, 'clasificacion': clasificacion, 'titulo': 'Editar Clasificación SERAP', 'accion': 'Guardar cambios'})


@login_required(login_url='login')
def cambiar_estado_clasificacion_serap(request, idclasificacion_serap):
	from .models import PatrimonioClasificacionSerap
	
	clasificacion = get_object_or_404(PatrimonioClasificacionSerap, idclasificacion_serap=idclasificacion_serap)
	
	if request.method == 'POST':
		clasificacion.activo = not clasificacion.activo
		clasificacion.save()
		estado = 'activada' if clasificacion.activo else 'desactivada'
		messages.success(request, f'Clasificación SERAP {estado} exitosamente.')
	
	return redirect('listar_clasificaciones_serap')


# ============== CATÁLOGOS DE PATRIMONIO: CLASIFICACIÓN CONTRALORÍA ==============

@login_required(login_url='login')
def listar_clasificaciones_contraloria(request):
	from .models import PatrimonioClasificacionContraloria
	
	search = request.GET.get('search', '').strip()
	clasificaciones = PatrimonioClasificacionContraloria.objects.all().order_by('nombre')
	
	if search:
		clasificaciones = clasificaciones.filter(Q(nombre__icontains=search) | Q(descripcion__icontains=search))
	
	return render(request, 'desarrollo/patrimonio/listar_clasificaciones_contraloria.html', {'clasificaciones': clasificaciones, 'search': search})


@login_required(login_url='login')
def crear_clasificacion_contraloria(request):
	from .forms import PatrimonioClasificacionContraloriaForm
	
	if request.method == 'POST':
		form = PatrimonioClasificacionContraloriaForm(request.POST)
		if form.is_valid():
			form.save()
			messages.success(request, 'Clasificación de Contraloría creada exitosamente.')
			return redirect('listar_clasificaciones_contraloria')
	else:
		form = PatrimonioClasificacionContraloriaForm()
	
	return render(request, 'desarrollo/patrimonio/form_clasificacion_contraloria.html', {'form': form, 'titulo': 'Nueva Clasificación de Contraloría', 'accion': 'Crear'})


@login_required(login_url='login')
def editar_clasificacion_contraloria(request, idclasificacion_contraloria):
	from .models import PatrimonioClasificacionContraloria
	from .forms import PatrimonioClasificacionContraloriaForm
	
	clasificacion = get_object_or_404(PatrimonioClasificacionContraloria, idclasificacion_contraloria=idclasificacion_contraloria)
	
	if request.method == 'POST':
		form = PatrimonioClasificacionContraloriaForm(request.POST, instance=clasificacion)
		if form.is_valid():
			form.save()
			messages.success(request, 'Clasificación de Contraloría actualizada exitosamente.')
			return redirect('listar_clasificaciones_contraloria')
	else:
		form = PatrimonioClasificacionContraloriaForm(instance=clasificacion)
	
	return render(request, 'desarrollo/patrimonio/form_clasificacion_contraloria.html', {'form': form, 'clasificacion': clasificacion, 'titulo': 'Editar Clasificación de Contraloría', 'accion': 'Guardar cambios'})


@login_required(login_url='login')
def cambiar_estado_clasificacion_contraloria(request, idclasificacion_contraloria):
	from .models import PatrimonioClasificacionContraloria
	
	clasificacion = get_object_or_404(PatrimonioClasificacionContraloria, idclasificacion_contraloria=idclasificacion_contraloria)
	
	if request.method == 'POST':
		clasificacion.activo = not clasificacion.activo
		clasificacion.save()
		estado = 'activada' if clasificacion.activo else 'desactivada'
		messages.success(request, f'Clasificación de Contraloría {estado} exitosamente.')
	
	return redirect('listar_clasificaciones_contraloria')

# ========== RESGUARDOS INTERNOS ==========

@login_required
def listar_resguardos(request):
	"""Lista todos los resguardos activos"""
	from .models import PatrimonioResguardo
	
	search = request.GET.get('search', '')
	filtro_estado = request.GET.get('estado', 'activos')  # activos, devueltos, todos
	
	resguardos = PatrimonioResguardo.objects.select_related('bien', 'empleado').all()
	
	# Filtrar por estado
	if filtro_estado == 'activos':
		resguardos = resguardos.filter(activo=True)
	elif filtro_estado == 'devueltos':
		resguardos = resguardos.filter(activo=False)
	
	# Búsqueda
	if search:
		resguardos = resguardos.filter(
			models.Q(bien__numero_inventario_itavu__icontains=search) |
			models.Q(bien__descripcion__icontains=search) |
			models.Q(empleado__nombre__icontains=search) |
			models.Q(empleado__apellido_paterno__icontains=search) |
			models.Q(empleado__apellido_materno__icontains=search)
		)
	
	context = {
		'resguardos': resguardos,
		'search': search,
		'filtro_estado': filtro_estado,
	}
	return render(request, 'desarrollo/patrimonio/listar_resguardos.html', context)


@login_required
def asignar_resguardo(request):
	"""Asigna un bien a un empleado (crear nuevo resguardo o transferir de un empleado a otro)"""
	from .forms import PatrimonioResguardoAsignacionForm
	from .models import PatrimonioResguardo
	from django.utils import timezone
	
	if request.method == 'POST':
		form = PatrimonioResguardoAsignacionForm(request.POST)
		if form.is_valid():
			bien = form.cleaned_data['bien']
			nuevo_empleado = form.cleaned_data['empleado']
			fecha_asignacion = form.cleaned_data['fecha_asignacion']
			observaciones = form.cleaned_data['observaciones_asignacion']
			
			# Verificar si el bien ya tiene un resguardo activo
			resguardo_actual = PatrimonioResguardo.objects.filter(bien=bien, activo=True).first()
			
			if resguardo_actual:
				# Es una TRANSFERENCIA/REASIGNACIÓN
				empleado_anterior = resguardo_actual.empleado
				
				# Cerrar el resguardo actual automáticamente
				resguardo_actual.activo = False
				resguardo_actual.fecha_devolucion = fecha_asignacion
				resguardo_actual.observaciones_devolucion = f"Transferido automáticamente a {nuevo_empleado.nombre} {nuevo_empleado.apellido_paterno}"
				resguardo_actual.usuario_devolucion = request.user.usuario if hasattr(request.user, 'usuario') else str(request.user)
				resguardo_actual.save()
				
				# Crear el nuevo resguardo
				nuevo_resguardo = PatrimonioResguardo.objects.create(
					bien=bien,
					empleado=nuevo_empleado,
					fecha_asignacion=fecha_asignacion,
					observaciones_asignacion=f"Transferido desde {empleado_anterior.nombre} {empleado_anterior.apellido_paterno}. {observaciones}",
					activo=True,
					usuario_asignacion=request.user.usuario if hasattr(request.user, 'usuario') else str(request.user)
				)
				
				messages.success(request, f'✓ Bien {bien.numero_inventario_itavu} transferido exitosamente de {empleado_anterior.nombre} {empleado_anterior.apellido_paterno} a {nuevo_empleado.nombre} {nuevo_empleado.apellido_paterno}.')
			else:
				# Es una ASIGNACIÓN NUEVA
				resguardo = form.save(commit=False)
				resguardo.activo = True
				resguardo.usuario_asignacion = request.user.usuario if hasattr(request.user, 'usuario') else str(request.user)
				resguardo.save()
				messages.success(request, f'Bien {resguardo.bien.numero_inventario_itavu} asignado exitosamente a {resguardo.empleado.nombre} {resguardo.empleado.apellido_paterno}.')
			
			return redirect('listar_resguardos')
	else:
		form = PatrimonioResguardoAsignacionForm()
	
	context = {
		'form': form,
		'titulo': 'Asignar o Transferir Resguardo',
		'accion': 'Asignar',
	}
	return render(request, 'desarrollo/patrimonio/form_resguardo_asignacion.html', context)


@login_required
def devolver_resguardo(request, idresguardo):
	"""Registra la devolución de un bien (termina el resguardo)"""
	from .models import PatrimonioResguardo
	from .forms import PatrimonioResguardoDevolucionForm
	
	resguardo = get_object_or_404(PatrimonioResguardo, idresguardo=idresguardo, activo=True)
	
	if request.method == 'POST':
		form = PatrimonioResguardoDevolucionForm(request.POST, instance=resguardo)
		if form.is_valid():
			resguardo = form.save(commit=False)
			resguardo.activo = False
			resguardo.usuario_devolucion = request.user.usuario if hasattr(request.user, 'usuario') else str(request.user)
			resguardo.save()
			messages.success(request, f'Devolución registrada exitosamente para el bien {resguardo.bien.numero_inventario_itavu}.')
			return redirect('listar_resguardos')
	else:
		form = PatrimonioResguardoDevolucionForm(instance=resguardo)
	
	context = {
		'form': form,
		'resguardo': resguardo,
		'titulo': 'Registrar Devolución',
		'accion': 'Registrar Devolución',
	}
	return render(request, 'desarrollo/patrimonio/form_resguardo_devolucion.html', context)


@login_required
def descargar_oficio_resguardo(request, idresguardo):
	"""Descarga el PDF del oficio asociado a un resguardo"""
	import os
	from .models import PatrimonioResguardo
	
	resguardo = get_object_or_404(PatrimonioResguardo, idresguardo=idresguardo)
	
	# Verificar que el usuario tenga acceso
	if not hasattr(request.user, 'usuario'):
		return redirect('login')
	
	# Verificar que el resguardo tenga un archivo de oficio
	if not resguardo.archivo_oficio:
		messages.error(request, 'Este resguardo no tiene un archivo de oficio asociado.')
		return redirect('listar_resguardos')
	
	# Obtener la ruta del archivo
	archivo_path = resguardo.archivo_oficio.path
	
	# Verificar que el archivo existe
	if not os.path.exists(archivo_path):
		messages.error(request, 'El archivo del oficio no se encuentra en el servidor.')
		return redirect('listar_resguardos')
	
	# Generar nombre de descarga: OF_[numero_oficio].pdf
	nombre_descarga = f"OF_{resguardo.numero_oficio.replace('/', '_')}.pdf"
	
	# Descargar el archivo
	with open(archivo_path, 'rb') as archivo:
		response = FileResponse(archivo, content_type='application/pdf')
		response['Content-Disposition'] = f'attachment; filename="{nombre_descarga}"'
		return response


@login_required
def verificar_resguardo_bien(request, idbien):
	"""API para verificar si un bien tiene resguardo activo"""
	from django.http import JsonResponse
	from .models import PatrimonioResguardo
	
	try:
		resguardo = PatrimonioResguardo.objects.filter(bien_id=idbien, activo=True).select_related('empleado').first()
		
		if resguardo:
			return JsonResponse({
				'tiene_resguardo': True,
				'empleado': f"{resguardo.empleado.nombre} {resguardo.empleado.apellido_paterno} {resguardo.empleado.apellido_materno}",
				'fecha_asignacion': resguardo.fecha_asignacion.strftime('%d/%m/%Y'),
			})
		else:
			return JsonResponse({'tiene_resguardo': False})
	except Exception as e:
		return JsonResponse({'error': str(e)}, status=400)


@login_required
def historial_resguardo_bien(request, idbien):
	"""Muestra el historial completo de resguardos de un bien"""
	from .models import PatrimonioBienesDelInstituto, PatrimonioResguardo
	
	bien = get_object_or_404(PatrimonioBienesDelInstituto, idbien=idbien)
	historial = PatrimonioResguardo.objects.filter(bien=bien).select_related('empleado').order_by('-fecha_asignacion')
	
	context = {
		'bien': bien,
		'historial': historial,
	}
	return render(request, 'desarrollo/patrimonio/historial_resguardo_bien.html', context)


@login_required
def historial_resguardo_empleado(request, idempleado):
	"""Muestra el historial completo de resguardos de un empleado"""
	from .models import PersonalEmpleados, PatrimonioResguardo
	
	empleado = get_object_or_404(PersonalEmpleados, idempleado=idempleado)
	historial = PatrimonioResguardo.objects.filter(empleado=empleado).select_related('bien').order_by('-fecha_asignacion')
	
	context = {
		'empleado': empleado,
		'historial': historial,
	}
	return render(request, 'desarrollo/patrimonio/historial_resguardo_empleado.html', context)


@login_required
def listar_entregas_departamento(request):
	"""Lista el historial de entregas-recepción de departamento"""
	from .models import PatrimonioEntregaDepartamento

	search = request.GET.get('search', '').strip()
	entregas = PatrimonioEntregaDepartamento.objects.select_related(
		'departamento',
		'empleado_saliente',
		'empleado_entrante'
	).all()

	if search:
		entregas = entregas.filter(
			Q(departamento__departamento__icontains=search) |
			Q(empleado_saliente__nombre__icontains=search) |
			Q(empleado_saliente__apellido_paterno__icontains=search) |
			Q(empleado_entrante__nombre__icontains=search) |
			Q(empleado_entrante__apellido_paterno__icontains=search)
		)

	context = {
		'entregas': entregas,
		'search': search,
	}
	return render(request, 'desarrollo/patrimonio/listar_entregas_departamento.html', context)


@login_required
def crear_entrega_departamento(request):
	"""Transfiere automáticamente todos los bienes activos de un empleado saliente a uno entrante"""
	from .forms import PatrimonioEntregaDepartamentoForm
	from .models import PatrimonioResguardo, PatrimonioEntregaDepartamentoDetalle

	if request.method == 'POST':
		form = PatrimonioEntregaDepartamentoForm(request.POST)
		if form.is_valid():
			empleado_saliente = form.cleaned_data['empleado_saliente']
			empleado_entrante = form.cleaned_data['empleado_entrante']
			fecha_entrega = form.cleaned_data['fecha_entrega']

			resguardos_activos = PatrimonioResguardo.objects.filter(
				empleado=empleado_saliente,
				activo=True
			).select_related('bien').order_by('bien__numero_inventario_itavu')

			if not resguardos_activos.exists():
				form.add_error('empleado_saliente', 'El empleado seleccionado no tiene bienes activos para transferir.')
			else:
				usuario_actual = request.user.usuario if hasattr(request.user, 'usuario') else str(request.user)

				with transaction.atomic():
					entrega = form.save(commit=False)
					entrega.departamento = empleado_saliente.iddepartamento
					entrega.usuario_registro = usuario_actual
					entrega.total_bienes = 0
					entrega.save()

					total_transferidos = 0
					for resguardo_actual in resguardos_activos:
						resguardo_actual.activo = False
						resguardo_actual.fecha_devolucion = fecha_entrega
						resguardo_actual.observaciones_devolucion = (
							f"Transferido por entrega-recepción departamental #{entrega.identrega} "
							f"a {empleado_entrante.nombre} {empleado_entrante.apellido_paterno}"
						)
						resguardo_actual.usuario_devolucion = usuario_actual
						resguardo_actual.save()

						nuevo_resguardo = PatrimonioResguardo.objects.create(
							bien=resguardo_actual.bien,
							empleado=empleado_entrante,
							fecha_asignacion=fecha_entrega,
							observaciones_asignacion=(
								f"Recibido por entrega-recepción departamental #{entrega.identrega} "
								f"desde {empleado_saliente.nombre} {empleado_saliente.apellido_paterno}."
							),
							activo=True,
							usuario_asignacion=usuario_actual
						)

						PatrimonioEntregaDepartamentoDetalle.objects.create(
							entrega=entrega,
							bien=resguardo_actual.bien,
							resguardo_anterior=resguardo_actual,
							resguardo_nuevo=nuevo_resguardo,
							fecha_transferencia=fecha_entrega
						)

						total_transferidos += 1

					entrega.total_bienes = total_transferidos
					entrega.save(update_fields=['total_bienes'])

				messages.success(
					request,
					f'✓ Entrega-recepción registrada. Se transfirieron {total_transferidos} bienes de {empleado_saliente.nombre} a {empleado_entrante.nombre}.'
				)
				return redirect('detalle_entrega_departamento', identrega=entrega.identrega)
	else:
		form = PatrimonioEntregaDepartamentoForm(initial={'fecha_entrega': timezone.localdate()})

	context = {
		'form': form,
		'titulo': 'Entrega-Recepción de Departamento',
		'accion': 'Registrar Entrega-Recepción',
	}
	return render(request, 'desarrollo/patrimonio/form_entrega_departamento.html', context)


@login_required
def detalle_entrega_departamento(request, identrega):
	"""Muestra el detalle de bienes transferidos en una entrega-recepción"""
	from .models import PatrimonioEntregaDepartamento

	entrega = get_object_or_404(
		PatrimonioEntregaDepartamento.objects.select_related(
			'departamento',
			'empleado_saliente',
			'empleado_entrante'
		).prefetch_related('detalles__bien'),
		identrega=identrega
	)

	context = {
		'entrega': entrega,
		'detalles': entrega.detalles.all(),
	}
	return render(request, 'desarrollo/patrimonio/detalle_entrega_departamento.html', context)


# ============== IMPORTACIÓN DE BIENES DESDE EXCEL ==============

@login_required(login_url='login')
def analizar_excel_bienes(request):
	"""Analiza un archivo Excel y retorna los nombres de las columnas"""
	
	# Verificar que el usuario sea cannoguzman
	user_ident = getattr(request.user, 'usuario', None) or getattr(request.user, 'username', None)
	if user_ident != 'cannoguzman':
		return JsonResponse({'success': False, 'error': 'Acceso denegado'}, status=403)
	
	if request.method != 'POST':
		return JsonResponse({'success': False, 'error': 'Método no permitido'}, status=405)
	
	try:
		import openpyxl
		
		archivo = request.FILES.get('archivo')
		if not archivo:
			return JsonResponse({'success': False, 'error': 'No se proporcionó archivo'})
		
		# Leer el archivo Excel
		workbook = openpyxl.load_workbook(archivo, data_only=True)
		worksheet = workbook.active
		
		# Obtener los nombres de las columnas (primera fila)
		columnas = []
		for cell in worksheet[1]:
			if cell.value:
				columnas.append(str(cell.value).strip())
		
		if not columnas:
			return JsonResponse({'success': False, 'error': 'El archivo no tiene encabezados'})
		
		return JsonResponse({
			'success': True,
			'campos': columnas
		})
	
	except Exception as e:
		return JsonResponse({'success': False, 'error': str(e)})


@login_required(login_url='login')
def importar_bienes_excel(request):
	"""Importa bienes desde un archivo Excel según el mapeo proporcionado"""
	
	# Verificar que el usuario sea cannoguzman
	user_ident = getattr(request.user, 'usuario', None) or getattr(request.user, 'username', None)
	if user_ident != 'cannoguzman':
		return JsonResponse({'success': False, 'error': 'Acceso denegado'}, status=403)
	
	if request.method != 'POST':
		return JsonResponse({'success': False, 'error': 'Método no permitido'}, status=405)
	
	try:
		import openpyxl
		from datetime import datetime
		from .models import PatrimonioBienesDelInstituto, PatrimonioMarcas
		
		archivo = request.FILES.get('archivo')
		mapeo_json = request.POST.get('mapeo', '{}')
		
		if not archivo:
			return JsonResponse({'success': False, 'error': 'No se proporcionó archivo'})
		
		# Parsear el mapeo
		mapeo = json.loads(mapeo_json)
		
		# Leer el archivo Excel
		workbook = openpyxl.load_workbook(archivo, data_only=True)
		worksheet = workbook.active
		
		# Obtener los índices de las columnas según el mapeo
		primera_fila = list(worksheet[1])
		indices_mapeo = {}
		
		for col_idx, cell in enumerate(primera_fila):
			col_nombre = str(cell.value).strip() if cell.value else None
			if col_nombre in mapeo and mapeo[col_nombre]:
				indices_mapeo[col_idx] = mapeo[col_nombre]
		
		importados = 0
		errores = 0
		
		# Procesar cada fila (comenzando desde la segunda fila)
		for fila_idx, fila in enumerate(worksheet.iter_rows(min_row=2, values_only=True), start=2):
			try:
				datos_bien = {}
				
				for col_idx, campo_sistema in indices_mapeo.items():
					valor = fila[col_idx] if col_idx < len(fila) else None
					if valor:
						datos_bien[campo_sistema] = str(valor).strip()
				
				# Validar que al menos tenga número de inventario ITAVU y descripción
				if 'numero_inventario_itavu' not in datos_bien or 'descripcion' not in datos_bien:
					errores += 1
					continue
				
				# Crear o actualizar el bien
				bien, creado = PatrimonioBienesDelInstituto.objects.update_or_create(
					numero_inventario_itavu=datos_bien.get('numero_inventario_itavu'),
					defaults={
						'numero_inventario_gobierno': datos_bien.get('numero_inventario_gobierno', ''),
						'descripcion': datos_bien.get('descripcion', ''),
						'modelo': datos_bien.get('modelo', ''),
						'serie': datos_bien.get('serie', ''),
						'numero_factura': datos_bien.get('numero_factura', ''),
						'proveedor': datos_bien.get('proveedor', ''),
						'observaciones': datos_bien.get('observaciones', ''),
						'activo': True,
					}
				)
				
				# Procesar campos con relaciones (marca, costo, fechas)
				if 'marca' in datos_bien:
					try:
						marca = PatrimonioMarcas.objects.get(nombre=datos_bien['marca'])
						bien.marca = marca
					except PatrimonioMarcas.DoesNotExist:
						pass
				
				if 'costo_articulo' in datos_bien:
					try:
						bien.costo_articulo = float(datos_bien['costo_articulo'])
					except ValueError:
						pass
				
				if 'fecha_factura' in datos_bien:
					try:
						fecha = datetime.strptime(datos_bien['fecha_factura'], '%d/%m/%Y').date()
						bien.fecha_factura = fecha
					except ValueError:
						pass
				
				bien.usuario_captura = request.user.username
				bien.save()
				
				importados += 1
			
			except Exception as e:
				errores += 1
				continue
		
		return JsonResponse({
			'success': True,
			'importados': importados,
			'errores': errores
		})
	
	except Exception as e:
		return JsonResponse({'success': False, 'error': str(e)})
# ---------- Importar y analizar para marcas ----------

@login_required(login_url='login')
def analizar_excel_marcas(request):
    """Analiza un archivo Excel y retorna los nombres de las columnas para marcas"""
    # Verificar que el usuario sea cannoguzman
    user_ident = getattr(request.user, 'usuario', None) or getattr(request.user, 'username', None)
    if user_ident != 'cannoguzman':
        return JsonResponse({'success': False, 'error': 'Acceso denegado'}, status=403)

    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Método no permitido'}, status=405)

    try:
        import openpyxl

        archivo = request.FILES.get('archivo')
        if not archivo:
            return JsonResponse({'success': False, 'error': 'No se proporcionó archivo'})

        # Leer el archivo Excel
        workbook = openpyxl.load_workbook(archivo, data_only=True)
        worksheet = workbook.active

        # Obtener los nombres de las columnas (primera fila)
        columnas = []
        for cell in worksheet[1]:
            if cell.value:
                columnas.append(str(cell.value).strip())

        if not columnas:
            return JsonResponse({'success': False, 'error': 'El archivo no tiene encabezados'})

        return JsonResponse({
            'success': True,
            'campos': columnas
        })

    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})


@login_required(login_url='login')
def importar_marcas_excel(request):
    """Importa marcas desde un archivo Excel según el mapeo proporcionado"""
    # Verificar que el usuario sea cannoguzman
    user_ident = getattr(request.user, 'usuario', None) or getattr(request.user, 'username', None)
    if user_ident != 'cannoguzman':
        return JsonResponse({'success': False, 'error': 'Acceso denegado'}, status=403)

    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Método no permitido'}, status=405)

    try:
        import openpyxl
        from .models import CatalogosMarcas

        archivo = request.FILES.get('archivo')
        mapeo_json = request.POST.get('mapeo', '{}')

        if not archivo:
            return JsonResponse({'success': False, 'error': 'No se proporcionó archivo'})

        # Parsear el mapeo
        mapeo = json.loads(mapeo_json)

        # Leer el archivo Excel
        workbook = openpyxl.load_workbook(archivo, data_only=True)
        worksheet = workbook.active

        # Obtener los �ndices de las columnas seg�n el mapeo
        primera_fila = list(worksheet[1])
        indices_mapeo = {}
        for col_idx, cell in enumerate(primera_fila):
            col_nombre = str(cell.value).strip() if cell.value else None
            if col_nombre in mapeo and mapeo[col_nombre]:
                indices_mapeo[col_idx] = mapeo[col_nombre]

        importados = 0
        errores = 0
        # Procesar cada fila
        for fila_idx, fila in enumerate(worksheet.iter_rows(min_row=2, values_only=True), start=2):
            try:
                datos = {}
                for col_idx, campo in indices_mapeo.items():
                    valor = fila[col_idx] if col_idx < len(fila) else None
                    if valor is not None:
                        datos[campo] = str(valor).strip()
                # requerir nombre
                if 'nombre' not in datos:
                    errores += 1
                    continue
                
                # determinar si usar ID o nombre como clave
                id_marca = datos.get('idmarca')
                if id_marca:
                    try:
                        id_marca = int(id_marca)
                        # intentar obtener por ID
                        marca = CatalogosMarcas.objects.filter(idmarca=id_marca).first()
                        if marca:
                            # actualizar registro existente
                            marca.nombre = datos.get('nombre')
                            marca.descripcion = datos.get('descripcion', '')
                            marca.activo = datos.get('activo','').lower() in ['1','true','si','sí','activo'] or \
                                         (datos.get('idestatus') == '0' if datos.get('idestatus') else True)
                            marca.save()
                        else:
                            # crear nuevo registro con ID específico
                            marca = CatalogosMarcas.objects.create(
                                idmarca=id_marca,
                                nombre=datos.get('nombre'),
                                descripcion=datos.get('descripcion', ''),
                                activo=datos.get('activo','').lower() in ['1','true','si','sí','activo'] or \
                                     (datos.get('idestatus') == '0' if datos.get('idestatus') else True)
                            )
                    except (ValueError, IntegrityError):
                        errores += 1
                        continue
                else:
                    # lógica original: crear o actualizar por nombre
                    marca, creado = CatalogosMarcas.objects.update_or_create(
                        nombre=datos.get('nombre'),
                        defaults={
                            'descripcion': datos.get('descripcion',''),
                            'activo': datos.get('activo','').lower() in ['1','true','si','sí','activo'] or
                                     (datos.get('idestatus') == '0' if datos.get('idestatus') else True)
                        }
                    )
                importados += 1
            except Exception:
                errores += 1
                continue

        return JsonResponse({'success': True, 'importados': importados, 'errores': errores})

    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})


# ---------- Importar y analizar para proveedores ----------

@login_required(login_url='login')
def analizar_excel_proveedores(request):
    """Analiza un archivo Excel y retorna los nombres de las columnas para proveedores"""
    # Verificar que el usuario sea cannoguzman
    user_ident = getattr(request.user, 'usuario', None) or getattr(request.user, 'username', None)
    if user_ident != 'cannoguzman':
        return JsonResponse({'success': False, 'error': 'Acceso denegado'}, status=403)

    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Método no permitido'}, status=405)

    try:
        import openpyxl

        archivo = request.FILES.get('archivo')
        if not archivo:
            return JsonResponse({'success': False, 'error': 'No se proporcionó archivo'})

        # Leer el archivo Excel
        workbook = openpyxl.load_workbook(archivo, data_only=True)
        worksheet = workbook.active

        # Obtener los nombres de las columnas (primera fila)
        columnas = []
        for cell in worksheet[1]:
            if cell.value:
                columnas.append(str(cell.value).strip())

        if not columnas:
            return JsonResponse({'success': False, 'error': 'El archivo no tiene encabezados'})

        return JsonResponse({
            'success': True,
            'campos': columnas
        })

    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})


@login_required(login_url='login')
def importar_proveedores_excel(request):
    """Importa proveedores desde un archivo Excel según el mapeo proporcionado"""
    # Verificar que el usuario sea cannoguzman
    user_ident = getattr(request.user, 'usuario', None) or getattr(request.user, 'username', None)
    if user_ident != 'cannoguzman':
        return JsonResponse({'success': False, 'error': 'Acceso denegado'}, status=403)

    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Método no permitido'}, status=405)

    try:
        import openpyxl
        from .models import PatrimonioProveedor

        archivo = request.FILES.get('archivo')
        mapeo_json = request.POST.get('mapeo', '{}')

        if not archivo:
            return JsonResponse({'success': False, 'error': 'No se proporcionó archivo'})

        # Parsear el mapeo
        mapeo = json.loads(mapeo_json)

        # Leer el archivo Excel
        workbook = openpyxl.load_workbook(archivo, data_only=True)
        worksheet = workbook.active

        # Obtener los índices de las columnas según el mapeo
        primera_fila = list(worksheet[1])
        indices_mapeo = {}
        for col_idx, cell in enumerate(primera_fila):
            col_nombre = str(cell.value).strip() if cell.value else None
            if col_nombre in mapeo and mapeo[col_nombre]:
                indices_mapeo[col_idx] = mapeo[col_nombre]

        importados = 0
        errores = 0
        # Procesar cada fila
        for fila_idx, fila in enumerate(worksheet.iter_rows(min_row=2, values_only=True), start=2):
            try:
                datos = {}
                for col_idx, campo in indices_mapeo.items():
                    valor = fila[col_idx] if col_idx < len(fila) else None
                    if valor is not None:
                        datos[campo] = str(valor).strip()

                # requerir nombre y que no esté vacío
                nombre_proveedor = datos.get('nombre', '').strip()
                if not nombre_proveedor:
                    errores += 1
                    continue

                # determinar si usar ID o nombre como clave
                id_proveedor = datos.get('idproveedor')
                if id_proveedor:
                    try:
                        id_proveedor = int(id_proveedor)
                        # intentar obtener por ID
                        proveedor = PatrimonioProveedor.objects.filter(idproveedor=id_proveedor).first()
                        if proveedor:
                            # actualizar registro existente
                            proveedor.nombre = nombre_proveedor
                            # Solo actualizar RFC si no está vacío o si es diferente
                            rfc_nuevo = datos.get('rfc', '').strip()
                            if rfc_nuevo and rfc_nuevo != proveedor.rfc:
                                # Verificar que el RFC no exista en otro registro
                                if PatrimonioProveedor.objects.filter(rfc=rfc_nuevo).exclude(idproveedor=id_proveedor).exists():
                                    errores += 1
                                    continue
                                proveedor.rfc = rfc_nuevo
                            proveedor.telefono = datos.get('telefono', None) or None
                            proveedor.correo = datos.get('correo', None) or None
                            proveedor.domicilio = datos.get('domicilio', None) or None
                            proveedor.persona_contacto = datos.get('persona_contacto', None) or None
                            proveedor.descripcion = datos.get('descripcion', None) or None
                            proveedor.activo = datos.get('activo','').lower() in ['1','true','si','sí','activo'] or \
                                             (datos.get('idestatus') == '0' if datos.get('idestatus') else True)
                            proveedor.save()
                        else:
                            # crear nuevo registro con ID específico
                            rfc_nuevo = datos.get('rfc', '').strip()
                            # Verificar unicidad de nombre y RFC
                            if PatrimonioProveedor.objects.filter(nombre=nombre_proveedor).exists():
                                errores += 1
                                continue
                            if rfc_nuevo and PatrimonioProveedor.objects.filter(rfc=rfc_nuevo).exists():
                                errores += 1
                                continue
                            
                            proveedor = PatrimonioProveedor.objects.create(
                                idproveedor=id_proveedor,
                                nombre=nombre_proveedor,
                                rfc=rfc_nuevo if rfc_nuevo else None,
                                telefono=datos.get('telefono', None) or None,
                                correo=datos.get('correo', None) or None,
                                domicilio=datos.get('domicilio', None) or None,
                                persona_contacto=datos.get('persona_contacto', None) or None,
                                descripcion=datos.get('descripcion', None) or None,
                                activo=datos.get('activo','').lower() in ['1','true','si','sí','activo'] or \
                                     (datos.get('idestatus') == '0' if datos.get('idestatus') else True)
                            )
                    except (ValueError, IntegrityError):
                        errores += 1
                        continue
                else:
                    # lógica original: crear o actualizar por nombre
                    if PatrimonioProveedor.objects.filter(nombre=nombre_proveedor).exists():
                        # actualizar existente
                        proveedor = PatrimonioProveedor.objects.get(nombre=nombre_proveedor)
                        # Solo actualizar RFC si no está vacío o si es diferente
                        rfc_nuevo = datos.get('rfc', '').strip()
                        if rfc_nuevo and rfc_nuevo != proveedor.rfc:
                            # Verificar que el RFC no exista en otro registro
                            if PatrimonioProveedor.objects.filter(rfc=rfc_nuevo).exclude(nombre=nombre_proveedor).exists():
                                errores += 1
                                continue
                            proveedor.rfc = rfc_nuevo
                        proveedor.telefono = datos.get('telefono', None) or None
                        proveedor.correo = datos.get('correo', None) or None
                        proveedor.domicilio = datos.get('domicilio', None) or None
                        proveedor.persona_contacto = datos.get('persona_contacto', None) or None
                        proveedor.descripcion = datos.get('descripcion', None) or None
                        proveedor.activo = datos.get('activo','').lower() in ['1','true','si','sí','activo'] or \
                                         (datos.get('idestatus') == '0' if datos.get('idestatus') else True)
                        proveedor.save()
                    else:
                        # crear nuevo
                        rfc_nuevo = datos.get('rfc', '').strip()
                        # Verificar unicidad de RFC solo si no está vacío
                        if rfc_nuevo and PatrimonioProveedor.objects.filter(rfc=rfc_nuevo).exists():
                            errores += 1
                            continue
                        
                        proveedor, creado = PatrimonioProveedor.objects.update_or_create(
                            nombre=nombre_proveedor,
                            defaults={
                                'rfc': rfc_nuevo if rfc_nuevo else None,
                                'telefono': datos.get('telefono', None) or None,
                                'correo': datos.get('correo', None) or None,
                                'domicilio': datos.get('domicilio', None) or None,
                                'persona_contacto': datos.get('persona_contacto', None) or None,
                                'descripcion': datos.get('descripcion', None) or None,
                                'activo': datos.get('activo','').lower() in ['1','true','si','sí','activo'] or
                                         (datos.get('idestatus') == '0' if datos.get('idestatus') else True)
                            }
                        )
                importados += 1
            except Exception as e:
                # Log the error for debugging
                import sys
                print(f"Error en fila {fila_idx}: {str(e)}", file=sys.stderr)
                errores += 1
                continue

        return JsonResponse({'success': True, 'importados': importados, 'errores': errores})

    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})

# ================================================================
#  MÓDULO: TICKET DE SERVICIO
# ================================================================

from .models import TicketServicio, TicketServicioArchivo, TicketServicioComentario
import mimetypes
import os


def _usuario_puede_ver_ticket(user, ticket):
    if not getattr(user, 'is_authenticated', False):
        return False
    if getattr(user, 'id_empleado', None) == ticket.emisor_id:
        return True
    if ticket.departamento_destino_id and getattr(user, 'iddepartamento_id', None) == ticket.departamento_destino_id:
        return True
    # Compatibilidad con tickets antiguos dirigidos a un receptor específico
    if ticket.receptor_id and getattr(user, 'id_empleado', None) == ticket.receptor_id:
        return True
    return _usuario_es_admin_sistema(user)


def _usuario_en_departamento_destino(user, ticket):
    return bool(
        ticket.departamento_destino_id and
        getattr(user, 'iddepartamento_id', None) == ticket.departamento_destino_id
    )


@login_required
def listar_tickets(request):
    """Lista tickets enviados por el usuario o recibidos por su departamento."""
    tab = request.GET.get('tab', 'todos')
    estado = request.GET.get('estado', '')
    prioridad = request.GET.get('prioridad', '')
    busqueda = request.GET.get('q', '').strip()

    q_recibidos = Q(pk__in=[])
    if request.user.iddepartamento_id:
        q_recibidos = Q(departamento_destino_id=request.user.iddepartamento_id)

    # Compatibilidad con registros legacy que aún usan receptor directo
    q_recibidos = q_recibidos | Q(receptor=request.user)

    qs_todos = TicketServicio.objects.select_related(
        'emisor', 'receptor', 'departamento_destino', 'atendido_por'
    ).filter(
        Q(emisor=request.user) | q_recibidos
    )

    if tab == 'enviados':
        qs = qs_todos.filter(emisor=request.user)
    elif tab == 'recibidos':
        qs = qs_todos.filter(q_recibidos)
    else:
        qs = qs_todos

    if estado:
        qs = qs.filter(estado=estado)
    if prioridad:
        qs = qs.filter(prioridad=prioridad)
    if busqueda:
        qs = qs.filter(
            Q(asunto__icontains=busqueda) |
            Q(descripcion__icontains=busqueda) |
            Q(emisor__nombre_completo__icontains=busqueda) |
            Q(receptor__nombre_completo__icontains=busqueda) |
            Q(departamento_destino__departamento__icontains=busqueda)
        )

    # Contadores para las pestañas
    conteo_todos      = qs_todos.count()
    conteo_enviados   = qs_todos.filter(emisor=request.user).count()
    conteo_recibidos  = qs_todos.filter(q_recibidos).count()
    conteo_pendientes = qs_todos.filter(
        q_recibidos, estado__in=['abierto', 'en_proceso', 'en_espera']
    ).count()

    return render(request, 'desarrollo/web/tickets/listar_tickets.html', {
        'tickets': qs.order_by('-fecha_creacion'),
        'tab': tab,
        'estado': estado,
        'prioridad': prioridad,
        'busqueda': busqueda,
        'conteo_todos': conteo_todos,
        'conteo_enviados': conteo_enviados,
        'conteo_recibidos': conteo_recibidos,
        'conteo_pendientes': conteo_pendientes,
        'estados': TicketServicio.ESTADO_CHOICES,
        'prioridades': TicketServicio.PRIORIDAD_CHOICES,
    })


@login_required
def crear_ticket(request):
    """Crear un nuevo ticket dirigido a un departamento con archivos adjuntos."""
    departamentos = PersonalDepartamento.objects.filter(activo=True).order_by('departamento')

    if request.method == 'POST':
        asunto          = request.POST.get('asunto', '').strip()
        descripcion     = request.POST.get('descripcion', '').strip()
        categoria       = request.POST.get('categoria', 'solicitud')
        prioridad       = request.POST.get('prioridad', 'normal')
        departamento_id = request.POST.get('departamento_destino')
        fecha_venc      = request.POST.get('fecha_vencimiento') or None
        archivos        = request.FILES.getlist('archivos')

        errores = []
        if not asunto:
            errores.append('El asunto es obligatorio.')
        if not descripcion:
            errores.append('La descripción es obligatoria.')
        if not departamento_id:
            errores.append('Debes seleccionar un departamento destino.')

        if not errores:
            try:
                departamento = PersonalDepartamento.objects.get(iddepartamento=departamento_id, activo=True)

                ticket = TicketServicio.objects.create(
                    asunto=asunto,
                    descripcion=descripcion,
                    categoria=categoria,
                    prioridad=prioridad,
                    emisor=request.user,
                    departamento_destino=departamento,
                    fecha_vencimiento=fecha_venc,
                )

                for f in archivos:
                    mime_type, _ = mimetypes.guess_type(f.name)
                    TicketServicioArchivo.objects.create(
                        ticket=ticket,
                        archivo=f,
                        nombre_original=f.name,
                        tamanio=f.size,
                        tipo_mime=mime_type or '',
                        subido_por=request.user,
                    )

                TicketServicioComentario.objects.create(
                    ticket=ticket,
                    autor=request.user,
                    mensaje=f'Ticket creado y enviado al departamento {departamento.departamento}.',
                    cambio_estado='abierto',
                )

                messages.success(request, f'Ticket #{ticket.id_ticket} creado exitosamente.')
                return redirect('ver_ticket', id_ticket=ticket.id_ticket)

            except PersonalDepartamento.DoesNotExist:
                errores.append('El departamento seleccionado no existe o está inactivo.')
            except Exception as e:
                errores.append(f'Error al crear el ticket: {str(e)}')

        for err in errores:
            messages.error(request, err)

    return render(request, 'desarrollo/web/tickets/crear_ticket.html', {
        'departamentos': departamentos,
        'categorias': TicketServicio.CATEGORIA_CHOICES,
        'prioridades': TicketServicio.PRIORIDAD_CHOICES,
        'today': timezone.localdate(),
    })


@login_required
def ver_ticket(request, id_ticket):
    """Detalle completo del ticket con actividad y archivos."""
    ticket = get_object_or_404(
        TicketServicio.objects.select_related('emisor', 'receptor', 'departamento_destino', 'atendido_por'),
        id_ticket=id_ticket
    )

    if not _usuario_puede_ver_ticket(request.user, ticket):
        messages.error(request, 'No tienes acceso a este ticket.')
        return redirect('listar_tickets')

    if request.method == 'POST':
        accion = request.POST.get('accion')

        if accion == 'comentar':
            mensaje  = request.POST.get('mensaje', '').strip()
            archivos = request.FILES.getlist('archivos')
            if mensaje:
                if _usuario_en_departamento_destino(request.user, ticket) and not ticket.atendido_por_id:
                    ticket.atendido_por = request.user
                    if ticket.estado == 'abierto':
                        ticket.estado = 'en_proceso'
                    ticket.save(update_fields=['atendido_por', 'estado', 'fecha_actualizacion'])

                TicketServicioComentario.objects.create(
                    ticket=ticket,
                    autor=request.user,
                    mensaje=mensaje,
                )
                for f in archivos:
                    mime_type, _ = mimetypes.guess_type(f.name)
                    TicketServicioArchivo.objects.create(
                        ticket=ticket,
                        archivo=f,
                        nombre_original=f.name,
                        tamanio=f.size,
                        tipo_mime=mime_type or '',
                        subido_por=request.user,
                    )
                messages.success(request, 'Comentario agregado.')
            else:
                messages.error(request, 'El comentario no puede estar vacío.')

        elif accion == 'tomar_ticket':
            if not _usuario_en_departamento_destino(request.user, ticket):
                messages.error(request, 'Solo personal del departamento destino puede tomar el ticket.')
            else:
                ticket.atendido_por = request.user
                if ticket.estado == 'abierto':
                    ticket.estado = 'en_proceso'
                ticket.save()
                TicketServicioComentario.objects.create(
                    ticket=ticket,
                    autor=request.user,
                    mensaje='Ticket tomado para atención.',
                )
                messages.success(request, 'Ticket asignado a tu atención.')

        elif accion == 'cambiar_estado':
            nuevo_estado    = request.POST.get('nuevo_estado')
            estados_validos = [s[0] for s in TicketServicio.ESTADO_CHOICES]
            if nuevo_estado in estados_validos:
                estado_anterior = ticket.get_estado_display()
                ticket.estado   = nuevo_estado
                if nuevo_estado == 'resuelto':
                    ticket.fecha_resolucion = timezone.now()
                ticket.save()
                nota = request.POST.get('nota_cambio', '').strip()
                msg  = nota if nota else f'Estado cambiado a: {ticket.get_estado_display()}'
                TicketServicioComentario.objects.create(
                    ticket=ticket,
                    autor=request.user,
                    mensaje=msg,
                    cambio_estado=estado_anterior,
                )
                messages.success(request, f'Estado actualizado a: {ticket.get_estado_display()}')
            else:
                messages.error(request, 'Estado no válido.')

        return redirect('ver_ticket', id_ticket=id_ticket)

    comentarios = ticket.comentarios.select_related('autor').order_by('fecha')
    archivos    = ticket.archivos.select_related('subido_por').order_by('fecha_subida')

    return render(request, 'desarrollo/web/tickets/ver_ticket.html', {
        'ticket': ticket,
        'comentarios': comentarios,
        'archivos': archivos,
        'estados': TicketServicio.ESTADO_CHOICES,
        'es_emisor': request.user.id_empleado == ticket.emisor.id_empleado,
        'es_receptor': _usuario_en_departamento_destino(request.user, ticket),
        'puede_tomar_ticket': _usuario_en_departamento_destino(request.user, ticket),
    })


@login_required
def descargar_archivo_ticket(request, id_archivo):
    """Descarga un archivo adjunto de un ticket."""
    archivo = get_object_or_404(TicketServicioArchivo, id_archivo=id_archivo)
    ticket  = archivo.ticket

    if not _usuario_puede_ver_ticket(request.user, ticket):
        messages.error(request, 'No tienes permiso para descargar este archivo.')
        return redirect('listar_tickets')

    try:
        response = FileResponse(
            archivo.archivo.open('rb'),
            as_attachment=True,
            filename=archivo.nombre_original,
        )
        return response
    except Exception:
        messages.error(request, 'No se pudo descargar el archivo.')
        return redirect('ver_ticket', id_ticket=ticket.id_ticket)


@login_required
def eliminar_archivo_ticket(request, id_archivo):
    """Elimina un archivo adjunto (solo el que lo subió)."""
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Método no permitido'}, status=405)

    archivo = get_object_or_404(TicketServicioArchivo, id_archivo=id_archivo)

    if archivo.subido_por and archivo.subido_por.id_empleado != request.user.id_empleado:
        return JsonResponse({'success': False, 'error': 'Sin permiso'}, status=403)

    try:
        if archivo.archivo and os.path.isfile(archivo.archivo.path):
            os.remove(archivo.archivo.path)
        archivo.delete()
        return JsonResponse({'success': True})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})


# ================================================================
#  MÓDULO: SERVICIO DE MANTENIMIENTO
# ================================================================

from .models import ConfiguracionSistema, TicketMantenimiento, TicketMantenimientoArchivo, TicketMantenimientoSeguimiento
from .forms import TicketMantenimientoCrearForm
from .email_utils import enviar_notificacion_mantenimiento


def _obtener_departamento_soporte_mantenimiento():
	config = ConfiguracionSistema.objects.first()
	departamento = getattr(config, 'departamento_soporte_mantenimiento', None)
	if departamento and departamento.activo:
		return departamento

	return PersonalDepartamento.objects.filter(
		activo=True,
		departamento__icontains='soporte'
	).order_by('departamento').first()


def _usuario_puede_ver_ticket_mantenimiento(user, ticket):
	if not getattr(user, 'is_authenticated', False):
		return False
	if getattr(user, 'id_empleado', None) == ticket.solicitante_id:
		return True
	if ticket.departamento_soporte_id and getattr(user, 'iddepartamento_id', None) == ticket.departamento_soporte_id:
		return True
	return _usuario_es_admin_sistema(user)


def _usuario_es_soporte_mantenimiento(user, ticket):
	return bool(
		ticket.departamento_soporte_id and
		getattr(user, 'iddepartamento_id', None) == ticket.departamento_soporte_id
	)


def _notificar_area_soporte(ticket, resumen):
	soporte_qs = PersonalEmpleados.objects.filter(
		activo=True,
		iddepartamento=ticket.departamento_soporte,
	).exclude(email='').values_list('email', 'id_empleado', 'nombre_completo')

	for email, id_empleado, nombre_completo in soporte_qs:
		enviar_notificacion_mantenimiento(
			ticket,
			email,
			'Nuevo ticket recibido',
			resumen,
			id_empleado=id_empleado,
			nombre_destinatario=nombre_completo,
		)


@login_required
def listar_tickets_mantenimiento(request):
	tab = request.GET.get('tab')
	estado = request.GET.get('estado', '')
	prioridad = request.GET.get('prioridad', '')
	busqueda = request.GET.get('q', '').strip()
	es_admin = _usuario_es_admin_sistema(request.user)

	q_recibidos = Q(pk__in=[])
	if es_admin:
		q_recibidos = Q()
		qs_todos = TicketMantenimiento.objects.select_related(
			'solicitante', 'departamento_solicitante', 'departamento_soporte', 'atendido_por'
		)
	else:
		if request.user.iddepartamento_id:
			q_recibidos = Q(departamento_soporte_id=request.user.iddepartamento_id)
		qs_todos = TicketMantenimiento.objects.select_related(
			'solicitante', 'departamento_solicitante', 'departamento_soporte', 'atendido_por'
		).filter(
			Q(solicitante=request.user) | q_recibidos
		)

	if not tab:
		es_soporte_con_tickets = bool(
			request.user.iddepartamento_id and qs_todos.filter(departamento_soporte_id=request.user.iddepartamento_id).exists()
		)
		tab = 'recibidos' if (es_admin or es_soporte_con_tickets) else 'todos'

	if tab == 'solicitados':
		qs = qs_todos.filter(solicitante=request.user)
	elif tab == 'recibidos':
		qs = qs_todos.filter(q_recibidos)
	else:
		qs = qs_todos

	if estado:
		qs = qs.filter(estado=estado)
	if prioridad:
		qs = qs.filter(prioridad=prioridad)
	if busqueda:
		qs = qs.filter(
			Q(asunto__icontains=busqueda) |
			Q(descripcion__icontains=busqueda) |
			Q(equipo__icontains=busqueda) |
			Q(numero_inventario__icontains=busqueda) |
			Q(solicitante__nombre_completo__icontains=busqueda) |
			Q(departamento_solicitante__departamento__icontains=busqueda)
		)

	return render(request, 'desarrollo/web/mantenimiento/listar_tickets_mantenimiento.html', {
		'tickets': qs.order_by('-fecha_creacion'),
		'tab': tab,
		'estado': estado,
		'prioridad': prioridad,
		'busqueda': busqueda,
		'conteo_todos': qs_todos.count(),
		'conteo_solicitados': qs_todos.filter(solicitante=request.user).count(),
		'conteo_recibidos': qs_todos.filter(q_recibidos).count(),
		'conteo_pendientes': qs_todos.filter(q_recibidos, estado__in=['abierto', 'asignado', 'en_revision', 'en_reparacion', 'espera_refaccion']).count(),
		'estados': TicketMantenimiento.ESTADO_CHOICES,
		'prioridades': TicketMantenimiento.PRIORIDAD_CHOICES,
		'es_soporte': es_admin or bool(request.user.iddepartamento_id and qs_todos.filter(departamento_soporte_id=request.user.iddepartamento_id).exists()),
	})


@login_required
def crear_ticket_mantenimiento(request):
	departamento_soporte = _obtener_departamento_soporte_mantenimiento()
	if not departamento_soporte:
		messages.error(request, 'No hay un departamento de soporte configurado. Solicita al administrador definirlo en Configuración del Sistema.')
		return redirect('listar_tickets_mantenimiento')

	if request.method == 'POST':
		form = TicketMantenimientoCrearForm(request.POST)
		archivos = request.FILES.getlist('archivos')
		if form.is_valid():
			ticket = form.save(commit=False)
			ticket.solicitante = request.user
			ticket.departamento_solicitante = request.user.iddepartamento
			ticket.departamento_soporte = departamento_soporte
			ticket.sla_horas_objetivo = TicketMantenimiento.SLA_HORAS_POR_PRIORIDAD.get(ticket.prioridad, 48)
			ticket.fecha_limite_sla = timezone.now() + timedelta(hours=ticket.sla_horas_objetivo)
			ticket.save()

			for f in archivos:
				mime_type, _ = mimetypes.guess_type(f.name)
				TicketMantenimientoArchivo.objects.create(
					ticket=ticket,
					archivo=f,
					nombre_original=f.name,
					tamanio=f.size,
					tipo_mime=mime_type or '',
					subido_por=request.user,
				)

			TicketMantenimientoSeguimiento.objects.create(
				ticket=ticket,
				autor=request.user,
				tipo='notificacion',
				mensaje=f'Ticket creado y enviado al departamento {departamento_soporte.departamento}.',
			)

			_notificar_area_soporte(ticket, 'Se registró una nueva solicitud de mantenimiento.')

			messages.success(request, f'Ticket de mantenimiento #{ticket.id_ticket_mantenimiento} creado exitosamente.')
			return redirect('ver_ticket_mantenimiento', id_ticket_mantenimiento=ticket.id_ticket_mantenimiento)
	else:
		form = TicketMantenimientoCrearForm()

	return render(request, 'desarrollo/web/mantenimiento/crear_ticket_mantenimiento.html', {
		'form': form,
		'departamento_soporte': departamento_soporte,
	})


@login_required
def ver_ticket_mantenimiento(request, id_ticket_mantenimiento):
	ticket = get_object_or_404(
		TicketMantenimiento.objects.select_related(
			'solicitante', 'departamento_solicitante', 'departamento_soporte', 'atendido_por'
		),
		id_ticket_mantenimiento=id_ticket_mantenimiento,
	)

	if not _usuario_puede_ver_ticket_mantenimiento(request.user, ticket):
		messages.error(request, 'No tienes acceso a este ticket de mantenimiento.')
		return redirect('listar_tickets_mantenimiento')

	es_admin = _usuario_es_admin_sistema(request.user)
	es_soporte = _usuario_es_soporte_mantenimiento(request.user, ticket)

	if request.method == 'POST':
		accion = request.POST.get('accion')

		if accion == 'comentar':
			mensaje = request.POST.get('mensaje', '').strip()
			archivos = request.FILES.getlist('archivos')
			notificar_solicitante = request.POST.get('notificar_solicitante') == 'on'

			if not mensaje:
				messages.error(request, 'El comentario no puede estar vacío.')
				return redirect('ver_ticket_mantenimiento', id_ticket_mantenimiento=id_ticket_mantenimiento)

			if es_soporte and not ticket.atendido_por_id:
				ticket.atendido_por = request.user
				if ticket.estado == 'abierto':
					ticket.estado = 'asignado'
				ticket.save(update_fields=['atendido_por', 'estado', 'fecha_actualizacion'])

			TicketMantenimientoSeguimiento.objects.create(
				ticket=ticket,
				autor=request.user,
				tipo='comentario',
				mensaje=mensaje,
				notificar_solicitante=notificar_solicitante,
			)

			for f in archivos:
				mime_type, _ = mimetypes.guess_type(f.name)
				TicketMantenimientoArchivo.objects.create(
					ticket=ticket,
					archivo=f,
					nombre_original=f.name,
					tamanio=f.size,
					tipo_mime=mime_type or '',
					subido_por=request.user,
				)

			if es_soporte and notificar_solicitante:
				enviar_notificacion_mantenimiento(ticket, ticket.solicitante.email, 'Nueva actualización', mensaje, id_empleado=ticket.solicitante_id)

			messages.success(request, 'Actualización agregada al ticket.')

		elif accion == 'tomar_ticket':
			if not es_soporte:
				messages.error(request, 'Solo el personal de soporte puede tomar este ticket.')
			else:
				ticket.atendido_por = request.user
				if ticket.estado == 'abierto':
					ticket.estado = 'asignado'
				ticket.save()
				TicketMantenimientoSeguimiento.objects.create(
					ticket=ticket,
					autor=request.user,
					tipo='estado',
					mensaje='El ticket fue tomado por personal de soporte.',
				)
				messages.success(request, 'Ticket asignado a tu atención.')

		elif accion == 'registrar_intervencion':
			if not es_soporte:
				messages.error(request, 'Solo el personal de soporte puede registrar intervenciones.')
			else:
				trabajo_realizado = request.POST.get('trabajo_realizado', '').strip()
				piezas_instaladas = request.POST.get('piezas_instaladas', '').strip()
				nuevo_estado = request.POST.get('nuevo_estado', ticket.estado)
				retirar_equipo = request.POST.get('equipo_retirado') == 'on'
				notificar_solicitante = request.POST.get('notificar_solicitante') == 'on'

				estado_anterior = ticket.get_estado_display()
				hubo_cambio_estado = nuevo_estado in dict(TicketMantenimiento.ESTADO_CHOICES) and nuevo_estado != ticket.estado
				ticket.atendido_por = request.user
				ticket.trabajo_realizado = trabajo_realizado or ticket.trabajo_realizado
				ticket.piezas_instaladas = piezas_instaladas or ticket.piezas_instaladas
				ticket.equipo_retirado = retirar_equipo
				if retirar_equipo:
					ticket.fecha_retiro_equipo = ticket.fecha_retiro_equipo or timezone.now()
				if nuevo_estado in dict(TicketMantenimiento.ESTADO_CHOICES):
					ticket.estado = nuevo_estado
				if nuevo_estado == 'resuelto':
					ticket.fecha_resolucion = timezone.now()
				if nuevo_estado in ['entregado', 'cerrado']:
					ticket.fecha_entrega_equipo = timezone.now()
					ticket.solucion_confirmada = True
				ticket.save()

				resumen = trabajo_realizado or 'Se registró una intervención técnica.'
				TicketMantenimientoSeguimiento.objects.create(
					ticket=ticket,
					autor=request.user,
					tipo='intervencion',
					mensaje=resumen,
					trabajo_realizado=trabajo_realizado,
					piezas_instaladas=piezas_instaladas,
					cambio_estado=estado_anterior if hubo_cambio_estado else '',
					notificar_solicitante=notificar_solicitante,
				)

				if notificar_solicitante:
					enviar_notificacion_mantenimiento(
						ticket,
						ticket.solicitante.email,
						'Intervención registrada',
						resumen,
						id_empleado=ticket.solicitante_id,
					)

				messages.success(request, 'Intervención registrada correctamente.')

		elif accion == 'cambiar_estado':
			if not es_soporte:
				messages.error(request, 'Solo el personal de soporte puede cambiar el estado del ticket.')
				return redirect('ver_ticket_mantenimiento', id_ticket_mantenimiento=id_ticket_mantenimiento)

			nuevo_estado = request.POST.get('nuevo_estado')
			nota = request.POST.get('nota_cambio', '').strip()
			notificar_solicitante = request.POST.get('notificar_solicitante') == 'on'

			if nuevo_estado in dict(TicketMantenimiento.ESTADO_CHOICES):
				estado_anterior = ticket.get_estado_display()
				hubo_cambio_estado = nuevo_estado != ticket.estado
				ticket.estado = nuevo_estado
				if nuevo_estado == 'resuelto':
					ticket.fecha_resolucion = timezone.now()
				if nuevo_estado in ['entregado', 'cerrado']:
					ticket.fecha_entrega_equipo = timezone.now()
					ticket.solucion_confirmada = True
				ticket.save()

				mensaje = nota or f'Estado actualizado a {ticket.get_estado_display()}.'
				TicketMantenimientoSeguimiento.objects.create(
					ticket=ticket,
					autor=request.user,
					tipo='estado',
					mensaje=mensaje,
					cambio_estado=estado_anterior if hubo_cambio_estado else '',
					notificar_solicitante=notificar_solicitante,
				)

				if notificar_solicitante and request.user.id_empleado != ticket.solicitante_id:
					enviar_notificacion_mantenimiento(ticket, ticket.solicitante.email, 'Cambio de estado', mensaje, id_empleado=ticket.solicitante_id)

				messages.success(request, f'Estado actualizado a {ticket.get_estado_display()}.')
			else:
				messages.error(request, 'Estado no válido.')

		return redirect('ver_ticket_mantenimiento', id_ticket_mantenimiento=id_ticket_mantenimiento)

	seguimientos = ticket.seguimientos.select_related('autor').order_by('fecha')
	archivos = ticket.archivos.select_related('subido_por').order_by('fecha_subida')

	return render(request, 'desarrollo/web/mantenimiento/ver_ticket_mantenimiento.html', {
		'ticket': ticket,
		'seguimientos': seguimientos,
		'archivos': archivos,
		'estados': TicketMantenimiento.ESTADO_CHOICES,
		'es_solicitante': request.user.id_empleado == ticket.solicitante_id,
		'es_soporte': es_soporte,
		'es_admin': es_admin,
		'puede_tomar_ticket': es_soporte and not ticket.atendido_por_id,
	})


@login_required
def descargar_archivo_mantenimiento(request, id_archivo):
	archivo = get_object_or_404(TicketMantenimientoArchivo, id_archivo=id_archivo)
	ticket = archivo.ticket

	if not _usuario_puede_ver_ticket_mantenimiento(request.user, ticket):
		messages.error(request, 'No tienes permiso para descargar este archivo.')
		return redirect('listar_tickets_mantenimiento')

	try:
		return FileResponse(
			archivo.archivo.open('rb'),
			as_attachment=True,
			filename=archivo.nombre_original,
		)
	except Exception:
		messages.error(request, 'No se pudo descargar el archivo.')
		return redirect('ver_ticket_mantenimiento', id_ticket_mantenimiento=ticket.id_ticket_mantenimiento)


@login_required
def eliminar_archivo_mantenimiento(request, id_archivo):
	if request.method != 'POST':
		return JsonResponse({'success': False, 'error': 'Método no permitido'}, status=405)

	archivo = get_object_or_404(TicketMantenimientoArchivo, id_archivo=id_archivo)

	if archivo.subido_por and archivo.subido_por.id_empleado != request.user.id_empleado and not _usuario_es_admin_sistema(request.user):
		return JsonResponse({'success': False, 'error': 'Sin permiso'}, status=403)

	try:
		if archivo.archivo and os.path.isfile(archivo.archivo.path):
			os.remove(archivo.archivo.path)
		archivo.delete()
		return JsonResponse({'success': True})
	except Exception as e:
		return JsonResponse({'success': False, 'error': str(e)})